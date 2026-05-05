"""
PerilCar ERP - Import diretto da Excel
Esegui questo script SEPARATAMENTE dal server:

  python import_excel.py DANEA_.xlsx

Importa direttamente nel database senza passare dal browser.
Funziona anche con file da 100MB+.
"""

import sys
import os
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

def main():
    if len(sys.argv) < 2:
        print("Uso: python import_excel.py <file.xlsx>")
        print("Esempio: python import_excel.py DANEA_.xlsx")
        input("Premi Invio per uscire...")
        sys.exit(1)

    filepath = sys.argv[1]

    # Se il percorso non è assoluto, cerca nella cartella corrente
    if not os.path.isabs(filepath):
        # Prova nella cartella corrente
        if os.path.exists(filepath):
            filepath = os.path.abspath(filepath)
        # Prova nel Desktop
        elif os.path.exists(os.path.join(os.path.expanduser("~"), "Desktop", filepath)):
            filepath = os.path.join(os.path.expanduser("~"), "Desktop", filepath)
        else:
            print(f"File non trovato: {filepath}")
            input("Premi Invio per uscire...")
            sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  PerilCar ERP — Import Excel")
    print(f"{'='*60}")
    print(f"  File: {filepath}")
    print(f"  Dimensione: {os.path.getsize(filepath)//1024} KB")
    print(f"{'='*60}\n")

    try:
        import openpyxl
    except ImportError:
        print("Installo openpyxl...")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    from core.config import ConfigManager
    from core.database import DatabaseManager

    cfg = ConfigManager()
    db  = DatabaseManager(cfg.get("db_path"))
    print(f"  Database: {cfg.get('db_path')}\n")

    # Mappa colonne Danea
    DANEA_MAP = {
        "cod.":             "codice",
        "descrizione":      "nome",
        "tipologia":        "tipologia",
        "categoria":        "categoria",
        "sottocategoria":   "sottocategoria",
        "cod. udm":         "cod_udm",
        "cod. iva":         "cod_iva",
        "listino 1":        "listino1",
        "listino 2":        "listino2",
        "listino 3":        "listino3",
        "note":             "note",
        "cod. a barre":     "cod_barre",
        "internet":         "internet",
        "produttore":       "marca",
        "extra 1":          "extra1",
        "extra 2":          "extra2",
        "extra 3":          "extra3",
        "extra 4":          "extra4",
        "cod. fornitore":   "cod_fornitore",
        "fornitore":        "fornitore",
        "cod. prod. forn.": "cod_prod_forn",
        "prezzo forn.":     "prezzo_forn",
        "note fornitura":   "note_fornitura",
        "ord. a multipli di": "ord_multipli",
        "gg. ordine":       "gg_ordine",
        "scorta min.":      "scorta_minima",
        "ubicazione":       "ubicazione",
        "q.tà giacenza":    "esistenza",
        "stato magazzino":  "stato_magazzino",
        "immagine":         "immagine_path",
        # alias generici
        "cmp":              "codice",
        "articolo":         "nome",
        "es":               "esistenza",
        "scorta minima":    "scorta_minima",
        "anno da":          "anno_da",
        "anno a":           "anno_a",
        "marca":            "marca",
        "modello":          "modello",
        "colore":           "colore",
    }

    print("  Apertura file Excel (attendere)...")
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    print(f"  Foglio: {ws.title}")

    # Leggi intestazioni
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h or "").strip().lower() for h in header_row]
    col_map = {}
    for i, h in enumerate(header):
        field = DANEA_MAP.get(h, h.replace(" ","_").replace(".","").replace("'",""))
        col_map[field] = i

    print(f"  Colonne rilevate: {len(header)}")
    print(f"  Avvio import...\n")

    def get(row, field, default=None):
        idx = col_map.get(field)
        if idx is None or idx >= len(row): return default
        v = row[idx]
        if v is None: return default
        import datetime
        if isinstance(v, (datetime.datetime, datetime.date)):
            return str(v.year)
        s = str(v).strip()
        return s if s not in ("None","nan","") else default

    def toint(row, field):
        try: v = get(row, field, "0"); return int(float(v)) if v else 0
        except: return 0

    def tofloat(row, field):
        try: v = get(row, field, "0"); return float(v) if v else 0.0
        except: return 0.0

    importati = 0; aggiornati = 0; saltati = 0; row_n = 1
    BATCH = 500
    ADMIN_ID = 1  # ID utente admin

    with db._write_lock:
        conn = db.get_connection()
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_n += 1
                codice = get(row, "codice")
                nome   = get(row, "nome")

                if not codice or not nome:
                    saltati += 1
                    continue

                scorta    = toint(row, "scorta_minima")
                esistenza = toint(row, "esistenza")

                campi = {
                    "nome":           nome,
                    "tipologia":      get(row,"tipologia"),
                    "categoria":      get(row,"categoria"),
                    "sottocategoria": get(row,"sottocategoria"),
                    "cod_udm":        get(row,"cod_udm"),
                    "cod_iva":        get(row,"cod_iva"),
                    "listino1":       tofloat(row,"listino1"),
                    "listino2":       tofloat(row,"listino2"),
                    "listino3":       tofloat(row,"listino3"),
                    "note":           get(row,"note"),
                    "cod_barre":      get(row,"cod_barre"),
                    "internet":       get(row,"internet"),
                    "marca":          get(row,"marca"),
                    "extra1":         get(row,"extra1"),
                    "extra2":         get(row,"extra2"),
                    "extra3":         get(row,"extra3"),
                    "extra4":         get(row,"extra4"),
                    "cod_fornitore":  get(row,"cod_fornitore"),
                    "fornitore":      get(row,"fornitore"),
                    "cod_prod_forn":  get(row,"cod_prod_forn"),
                    "prezzo_forn":    tofloat(row,"prezzo_forn"),
                    "note_fornitura": get(row,"note_fornitura"),
                    "ord_multipli":   toint(row,"ord_multipli"),
                    "gg_ordine":      toint(row,"gg_ordine"),
                    "scorta_minima":  scorta,
                    "ubicazione":     get(row,"ubicazione"),
                    "stato_magazzino":get(row,"stato_magazzino"),
                }

                existing = conn.execute(
                    "SELECT id FROM componenti WHERE codice=? AND eliminato=0",
                    (codice,)).fetchone()

                if existing:
                    comp_id = existing[0]
                    sets = ", ".join(f"{k}=?" for k in campi)
                    conn.execute(
                        f"UPDATE componenti SET {sets}, modificato_il=datetime('now') WHERE id=?",
                        list(campi.values()) + [comp_id])
                    aggiornati += 1
                else:
                    campi["codice"]     = codice
                    campi["pubblicato"] = 0
                    campi["creato_da"]  = ADMIN_ID
                    cols_str     = ", ".join(campi.keys())
                    placeholders = ", ".join(["?"] * len(campi))
                    cur = conn.execute(
                        f"INSERT INTO componenti ({cols_str}) VALUES ({placeholders})",
                        list(campi.values()))
                    comp_id = cur.lastrowid
                    conn.execute(
                        "INSERT OR IGNORE INTO magazzino(componente_id,scorta_minima) VALUES(?,?)",
                        (comp_id, scorta))
                    importati += 1

                # Carico iniziale
                if esistenza > 0:
                    already = conn.execute(
                        "SELECT id FROM movimenti_magazzino WHERE componente_id=? AND riferimento='Import Excel' LIMIT 1",
                        (comp_id,)).fetchone()
                    if not already:
                        conn.execute(
                            """INSERT INTO movimenti_magazzino
                               (componente_id,tipo,quantita,quantita_prima,
                                quantita_dopo,riferimento,utente_id)
                               VALUES(?,?,?,?,?,?,?)""",
                            (comp_id,"carico",esistenza,0,esistenza,
                             "Import Excel", ADMIN_ID))

                # Progresso ogni 500 righe
                if row_n % BATCH == 0:
                    conn.commit()
                    tot_proc = importati + aggiornati
                    print(f"  Riga {row_n:>6} — Nuovi: {importati:>6}  Aggiornati: {aggiornati:>6}  Saltati: {saltati:>4}")

            conn.commit()
            wb.close()

        except Exception as e:
            conn.rollback()
            print(f"\n  ERRORE alla riga {row_n}: {e}")
            input("Premi Invio per uscire...")
            sys.exit(1)
        finally:
            conn.close()

    print(f"\n{'='*60}")
    print(f"  ✅ IMPORT COMPLETATO!")
    print(f"  Nuovi componenti:    {importati:>6}")
    print(f"  Aggiornati:          {aggiornati:>6}")
    print(f"  Saltati (no codice): {saltati:>6}")
    print(f"  Totale righe lette:  {row_n-1:>6}")
    print(f"{'='*60}")
    print(f"\n  Ora puoi aprire il browser su http://localhost:5000")
    input("\n  Premi Invio per uscire...")

if __name__ == "__main__":
    main()
