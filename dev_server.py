"""
PerilCar ERP — Dev Server v1.1
Flask + SocketIO con hot reload automatico.
"""
import sys, os, time, threading, logging, io, socket
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

# ── Configurazione percorsi (locale o NAS) ────────────────────────────
def _carica_config():
    """Legge config/settings.json se esiste, altrimenti usa percorsi locali."""
    cfg_path = ROOT / "config" / "settings.json"
    if cfg_path.exists():
        try:
            import json as _json
            return _json.loads(cfg_path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[WARN] Errore lettura settings.json: {e}")
    return {}

_cfg = _carica_config()
_nas = _cfg.get("nas", {})

# Percorso DB magazzino
if _nas.get("abilitato") and _nas.get("percorso_db"):
    _db_dir = Path(_nas["percorso_db"])
    try:
        # Verifica NAS raggiungibile in max 2 secondi
        import socket as _sock
        _nas_ip = _nas.get("ip", "")
        if _nas_ip:
            _s = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
            _s.settimeout(2)
            _reachable = _s.connect_ex((_nas_ip, 445)) == 0
            _s.close()
        else:
            _reachable = True
        if not _reachable:
            raise Exception(f"NAS {_nas_ip} non raggiungibile")
        _db_dir.mkdir(parents=True, exist_ok=True)
        _perilcar_db_path = _db_dir / "perilcar.db"
        _dem_db_path      = _db_dir / "demolizioni.db"
        print(f"[NAS] DB su: {_db_dir}")
    except Exception as e:
        print(f"[WARN] NAS non raggiungibile ({e}), uso DB locale")
        _perilcar_db_path = ROOT / "db" / "perilcar.db"
        _dem_db_path      = ROOT / "db" / "demolizioni.db"
else:
    _perilcar_db_path = ROOT / "db" / "perilcar.db"
    _dem_db_path      = ROOT / "db" / "demolizioni.db"

# Percorso uploads (foto)
if _nas.get("abilitato") and _nas.get("percorso_uploads"):
    UPLOAD_FOLDER = Path(_nas["percorso_uploads"])
    try:
        UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
        print(f"[NAS] Uploads su: {UPLOAD_FOLDER}")
    except Exception as e:
        print(f"[WARN] NAS uploads non raggiungibile ({e}), uso locale")
        UPLOAD_FOLDER = ROOT / "web" / "static" / "uploads"
        UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
else:
    UPLOAD_FOLDER = ROOT / "web" / "static" / "uploads"
    UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(name)-20s  %(levelname)s  %(message)s")
log = logging.getLogger("perilcar.dev")

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, Response
from flask_socketio import SocketIO, join_room, leave_room

try:
    import orjson as _orjson
    def _fast_json(obj):
        return Response(_orjson.dumps(obj), mimetype="application/json")
except ImportError:
    _fast_json = jsonify

app = Flask(__name__, template_folder="web/templates", static_folder="web/static")
_sk = _cfg.get("flask_secret_key", "")
if not _sk or len(_sk) < 32:
    import secrets as _sec, json as _j
    _sk = _sec.token_hex(32)
    _cfg["flask_secret_key"] = _sk
    _sk_path = ROOT / "config" / "settings.json"
    _sk_path.parent.mkdir(parents=True, exist_ok=True)
    _sk_path.write_text(_j.dumps(_cfg, indent=2, ensure_ascii=False), encoding="utf-8")
app.secret_key = _sk

@app.route('/favicon.ico')
def favicon():
    return '', 204  # No content - evita errori 404

app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024

@app.after_request
def no_cache(response):
    """Disabilita cache browser per tutti i file durante sviluppo."""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

from core.config import ConfigManager
from core.database import DatabaseManager

cfg = ConfigManager()
db  = DatabaseManager(cfg.get("db_path"))
log.info(f"DB magazzino: {cfg.get('db_path')}")

# ── Mapping colonne DB (schema Danea 18 col) ────────────────────────
_DB_COLS = {
    "cmp": "codice", "articolo": "descrizione", "nota": "note",
    "eliminato": "eliminato", "aggiornato_il": "aggiornato_il",
    "all_cols": [], "tabella_movimenti": "movimenti_magazzino"
}

def _auto_fix_view():
    """Ricrea v_giacenza con schema Danea fisso."""
    global _DB_COLS
    try:
        _raw = db.get_connection()
        cols = [r[1] for r in _raw.execute("PRAGMA table_info(componenti)").fetchall()]
        if not cols:
            _raw.close(); return

        tables = [r[0] for r in _raw.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        tab_mov = "movimenti_magazzino" if "movimenti_magazzino" in tables else "movimenti"

        _DB_COLS = {
            "cmp": "codice", "articolo": "descrizione", "nota": "note",
            "eliminato": "eliminato", "aggiornato_il": "aggiornato_il",
            "all_cols": cols, "tabella_movimenti": tab_mov,
        }
        log.info(f"DB schema: cmp=codice, art=descrizione, mov={tab_mov}")

        # Migrazione: aggiungi colonna giacenza se non esiste e popolala dai movimenti
        if "giacenza" not in cols:
            _raw.execute("ALTER TABLE componenti ADD COLUMN giacenza INTEGER NOT NULL DEFAULT 0")
            _raw.execute(f"""
                UPDATE componenti SET giacenza = COALESCE((
                    SELECT SUM(CASE
                        WHEN tipo IN ('carico','inventario') THEN  quantita
                        WHEN tipo = 'scarico'               THEN -quantita
                        WHEN tipo = 'rettifica'             THEN  quantita
                        ELSE 0 END)
                    FROM {tab_mov} WHERE componente_id = componenti.id
                ), 0)
            """)
            _raw.commit()
            cols.append("giacenza")
            log.info("Colonna giacenza aggiunta e popolata")

        def opt(c, alias=None):
            a = alias or c
            return f"c.{c} AS {a}" if c in cols else f"NULL AS {a}"

        _raw.execute("DROP VIEW IF EXISTS v_giacenza")
        _raw.execute(f"""
            CREATE VIEW v_giacenza AS
            SELECT
                c.id              AS componente_id,
                c.codice          AS cmp,
                c.descrizione     AS articolo,
                {opt('produttore','marca')},
                {opt('modello')},
                {opt('udm','cod_udm')},
                {opt('anno','anno_da')},
                {opt('cod_prod_forn')},
                {opt('alimentazione','carburante')},
                {opt('colore')},
                {opt('note','nota')},
                {opt('cilindrata')},
                {opt('ubicazione')},
                {opt('cod_barre')},
                {opt('extra3')},
                {opt('cod_fornitore')},
                {opt('fornitore')},
                {opt('immagine','immagine_path')},
                c.eliminato,
                c.aggiornato_il,
                c.giacenza        AS esistenza
            FROM componenti c WHERE c.eliminato=0
        """)
        _raw.commit()
        _raw.close()
        log.info("v_giacenza OK")
    except Exception as e:
        log.warning(f"Auto-fix view: {e}")

_auto_fix_view()

# Aggiungi indici per velocizzare le query
def _crea_indici():
    try:
        _raw = db.get_connection()
        tab_mov = _DB_COLS.get("tabella_movimenti", "movimenti_magazzino")
        col_cmp = _DB_COLS.get("cmp", "codice")
        _raw.execute(f"CREATE INDEX IF NOT EXISTS idx_mov_comp ON {tab_mov}(componente_id)")
        _raw.execute(f"CREATE INDEX IF NOT EXISTS idx_comp_art ON componenti({col_cmp})")
        _raw.execute("CREATE INDEX IF NOT EXISTS idx_comp_desc    ON componenti(descrizione)")
        _raw.execute("CREATE INDEX IF NOT EXISTS idx_comp_giacenza ON componenti(giacenza)")
        _raw.commit()
        _raw.close()
        log.info("Indici DB creati OK")
    except Exception as e:
        log.warning(f"Indici: {e}")

_crea_indici()

# DB separato per demolizioni
import os as _os
# Percorso demolizioni.db (NAS o locale, coerente con perilcar.db)
_dem_path = str(_dem_db_path) if '_dem_db_path' in dir() else \
            _os.path.join(_os.path.dirname(cfg.get("db_path")), "demolizioni.db")

class _DemDB:
    """Connessione SQLite semplice per DB demolizioni."""
    def __init__(self, path):
        self.path = path
        self._lock = __import__('threading').Lock()
        self._init()
    
    def _init(self):
        import sqlite3 as _sq
        conn = _sq.connect(self.path, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS anagrafiche (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nominativo TEXT NOT NULL,
            cognome TEXT, nome TEXT, cf_piva TEXT, sesso TEXT, tipo_societa TEXT,
            tipo TEXT DEFAULT 'privato', data_nascita TEXT, luogo_nascita TEXT,
            prov_nascita TEXT, comune TEXT, provincia TEXT, via TEXT, civico TEXT,
            cap TEXT, tipo_doc TEXT, num_doc TEXT, data_doc TEXT, rilasciato_da TEXT,
            telefono TEXT, cellulare TEXT, fax TEXT, email TEXT, indirizzo TEXT,
            note TEXT, creato_il TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS veicoli (
            id INTEGER PRIMARY KEY AUTOINCREMENT, targa TEXT, telaio TEXT,
            classe TEXT, marca TEXT, modello TEXT, anno_immatricolazione TEXT,
            num_motore TEXT, colore TEXT, note TEXT, creato_il TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS demolizioni (
            id INTEGER PRIMARY KEY AUTOINCREMENT, data_presa_in_carico TEXT,
            ora_presa_in_carico TEXT, reg_demolitori TEXT, pag_reg TEXT,
            veicolo_id INTEGER, proprietario_id INTEGER, detentore_id INTEGER,
            ufficio_provinciale TEXT, targhe_consegnate INTEGER DEFAULT 0,
            carta_circolazione INTEGER DEFAULT 0, concessionaria TEXT,
            peso_effettivo_kg REAL, peso_netto_kg REAL, modalita_radiazione TEXT,
            num_albatros TEXT, certificato_id TEXT, note TEXT,
            creato_da INTEGER, primo_trattamento INTEGER DEFAULT 0,
            creato_il TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS ricambi_sottratti (
            id INTEGER PRIMARY KEY AUTOINCREMENT, demolizione_id INTEGER NOT NULL,
            componente_id INTEGER, peso_kg REAL, note TEXT, pezzo_nome TEXT,
            creato_il TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS voci_tendine (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria TEXT NOT NULL,
            valore TEXT NOT NULL,
            ordine INTEGER DEFAULT 0,
            creato_il TEXT DEFAULT (datetime('now')),
            UNIQUE(categoria, valore)
        );
        -- Voci default per ogni categoria
        INSERT OR IGNORE INTO voci_tendine(categoria,valore,ordine) VALUES
            ('modalita','Cancellazione al PRA',1),
            ('modalita','Solo presa in carico',2),
            ('modalita','Radiazione targa e ciclomotore',3),
            ('modalita','Solo radiazione ciclomotore',4),
            ('classe','Autovettura',1),('classe','Motoveicolo',2),('classe','Autocarro',3),
            ('classe','Rimorchio',4),('classe','Macchina agricola',5),('classe','Altro',6),
            ('concessionaria','',1);

        CREATE TABLE IF NOT EXISTS schede_demolizione (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dem_ids TEXT NOT NULL,
            data_trattamento TEXT,
            peso_eff_tot REAL, peso_ricambi_tot REAL, peso_netto_tot REAL,
            righe_json TEXT,
            creato_da INTEGER, creato_il TEXT DEFAULT (datetime('now')),
            modificato_il TEXT
        );
        """)
        # Migrazione colonne opzionali
        cols = [r[1] for r in conn.execute("PRAGMA table_info(demolizioni)").fetchall()]
        if 'primo_trattamento' not in cols:
            conn.execute("ALTER TABLE demolizioni ADD COLUMN primo_trattamento INTEGER DEFAULT 0")
            conn.commit()
        conn.commit()
        conn.close()
    
    def conn(self):
        import sqlite3 as _sq
        c = _sq.connect(self.path, timeout=30)
        c.row_factory = _sq.Row
        c.execute("PRAGMA journal_mode=WAL")
        return c
    
    def all(self, sql, params=()):
        with self._lock:
            c = self.conn()
            try:
                rows = c.execute(sql, params).fetchall()
                return [dict(r) for r in rows]
            finally:
                c.close()
    
    def one(self, sql, params=()):
        with self._lock:
            c = self.conn()
            try:
                row = c.execute(sql, params).fetchone()
                return dict(row) if row else None
            finally:
                c.close()
    
    def run(self, sql, params=()):
        with self._lock:
            c = self.conn()
            try:
                cur = c.execute(sql, params)
                c.commit()
                return cur.lastrowid
            finally:
                c.close()

dem = _DemDB(_dem_path)
log.info(f"DB demolizioni: {_dem_path}")

# Pulizia tabelle residue nel magazzino DB
try:
    import sqlite3 as _sq3
    _cp = _sq3.connect(cfg.get("db_path"), timeout=10)
    _cp.execute("PRAGMA journal_mode=WAL")
    for _tbak in ['veicoli_bak','veicoli_old','veicoli_tmp']:
        if _cp.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (_tbak,)).fetchone():
            _cp.execute("DROP TABLE IF EXISTS ["+_tbak+"]")
            log.info(f"Rimossa tabella residua: {_tbak}")
    _cp.commit()
    _cp.close()
except Exception as _ce:
    log.warning(f"Pulizia avvio: {_ce}")


# ── Crea/migra tabelle inventario ─────────────────────────────────────────────
try:
    import sqlite3 as _sq_inv
    _inv_conn = _sq_inv.connect(cfg.get("db_path"), timeout=10)
    _inv_conn.executescript("""
    CREATE TABLE IF NOT EXISTS sessioni_inventario (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        categoria TEXT,
        filtro_json TEXT,
        stato TEXT DEFAULT 'attiva',
        totale_pezzi INTEGER DEFAULT 0,
        confermati INTEGER DEFAULT 0,
        mancanti INTEGER DEFAULT 0,
        sospesi INTEGER DEFAULT 0,
        creato_da INTEGER,
        creato_il TEXT DEFAULT (datetime('now')),
        chiuso_il TEXT
    );
    CREATE TABLE IF NOT EXISTS inventario_righe (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sessione_id INTEGER NOT NULL,
        componente_id INTEGER NOT NULL,
        stato TEXT DEFAULT 'sospeso',
        qty_attesa INTEGER DEFAULT 0,
        qty_trovata INTEGER,
        foto_url TEXT,
        note TEXT,
        ordine INTEGER DEFAULT 0,
        aggiornato_il TEXT,
        FOREIGN KEY(sessione_id) REFERENCES sessioni_inventario(id)
    );
    CREATE INDEX IF NOT EXISTS idx_inv_sessione ON inventario_righe(sessione_id);
    CREATE INDEX IF NOT EXISTS idx_inv_stato ON inventario_righe(sessione_id, stato);
    """)
    # Migrazione: aggiunge colonne mancanti a sessioni_inventario se DB vecchio
    _inv_existing = {r[1] for r in _inv_conn.execute("PRAGMA table_info(sessioni_inventario)").fetchall()}
    for _col, _defn in [
        ("categoria",     "TEXT"),
        ("filtro_json",   "TEXT"),
        ("totale_pezzi",  "INTEGER DEFAULT 0"),
        ("confermati",    "INTEGER DEFAULT 0"),
        ("mancanti",      "INTEGER DEFAULT 0"),
        ("sospesi",       "INTEGER DEFAULT 0"),
        ("creato_da",     "INTEGER"),
        ("chiuso_il",     "TEXT"),
    ]:
        if _col not in _inv_existing:
            _inv_conn.execute(f"ALTER TABLE sessioni_inventario ADD COLUMN {_col} {_defn}")
    # Se esiste sessioni_inventario_righe (schema vecchio), copia i dati in inventario_righe
    _old_tables = {r[0] for r in _inv_conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    if "sessioni_inventario_righe" in _old_tables and "inventario_righe" in _old_tables:
        _cnt = _inv_conn.execute("SELECT COUNT(*) FROM inventario_righe").fetchone()[0]
        if _cnt == 0:
            _inv_conn.execute("""
                INSERT OR IGNORE INTO inventario_righe
                    (sessione_id,componente_id,stato,qty_attesa,qty_trovata,note,ordine,aggiornato_il)
                SELECT sessione_id,componente_id,
                    COALESCE(stato,'sospeso'),
                    COALESCE(qty_attesa,0),qty_trovata,
                    note,COALESCE(ordine,0),aggiornato_il
                FROM sessioni_inventario_righe
            """)
    _inv_conn.commit()
    _inv_conn.close()
    log.info("Tabelle inventario pronte")
except Exception as _e:
    log.warning(f"Tabelle inventario: {_e}")

# ── Scrittura sicura: tutte le operazioni passano da qui ─────────────────────
def db_write(statements: list):
    """Esegue più statement in una singola transazione atomica."""
    with db._write_lock:
        conn = db.get_connection()
        try:
            for sql, params in statements:
                conn.execute(sql, params)
            conn.commit()
            conn.close()
        except Exception as e:
            conn.rollback()
            raise

def require_login(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            # API → 401 JSON; pagine → redirect login
            if request.path.startswith('/api/'):
                return jsonify({"ok": False, "msg": "Non autenticato"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return wrapper

def cu():
    return session.get("user", {})

# ══════════════════════════════════════════════════════════════════════════════
# PAGINE
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@require_login
def index():
    if session.get("user"):
        return redirect("/dashboard")
    return redirect("/login")

@app.route("/login")
def login_page():
    return render_template("login.html")

@app.route("/dashboard")
@require_login
def dashboard():
    return render_template("dashboard.html", user=session["user"])

@app.route("/magazzino")
@require_login
def magazzino():
    return render_template("magazzino.html", user=session["user"])

@app.route("/storico")
@require_login
def storico():
    return render_template("storico.html", user=session["user"])

# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/login", methods=["POST"])
def api_login():
    import hashlib
    d = request.json or {}
    pwd_hash = hashlib.sha256(d.get("password","").encode()).hexdigest()
    user = db.fetchone("SELECT * FROM utenti WHERE username=?",
                       (d.get("username","").strip(),))
    if not user:
        return jsonify({"ok": False, "msg": "Utente non trovato"}), 401
    if user.get("eliminato", 0) == 1:
        return jsonify({"ok": False, "msg": "Account eliminato"}), 403
    if user.get("attivo", 1) == 0:
        return jsonify({"ok": False, "msg": "Account disabilitato"}), 403
    if user["password_hash"] != pwd_hash:
        return jsonify({"ok": False, "msg": "Password errata"}), 401
    session["user"] = {"id": user["id"], "username": user["username"],
                       "ruolo": user["ruolo"], "nome": (user.get("nome_completo") or user["username"])}
    try:
        db_write([("INSERT INTO log_operazioni(utente_id,username,modulo,azione) VALUES(?,?,?,?)",
                   (user["id"], user["username"], "AUTH", "LOGIN"))])
    except: pass
    # Redirect: mobile -> inventario-mobile, ?next= -> pagina richiesta, default -> dashboard
    ua     = request.headers.get("User-Agent","").lower()
    is_mob = any(x in ua for x in ["mobile","android","iphone","ipad","mobi"])
    nxt    = (request.json or {}).get("next","")
    if nxt:
        redir = nxt
    elif is_mob:
        redir = "/inventario-mobile"
    else:
        redir = "/dashboard"
    return jsonify({"ok": True, "user": session["user"], "redirect": redir})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    u = cu()
    if u:
        db_write([("INSERT INTO log_operazioni(utente_id,username,modulo,azione) VALUES(?,?,?,?)",
                   (u.get("id"), u.get("username"), "AUTH", "LOGOUT"))])
    session.clear()
    return jsonify({"ok": True})

# ══════════════════════════════════════════════════════════════════════════════
# COMPONENTI
# ══════════════════════════════════════════════════════════════════════════════

# Cache componenti
_comp_cache = {"ts": 0, "data": None, "total": 0}
_COMP_CACHE_TTL = 120

@app.route("/api/magazzino/componenti")
@require_login
# Cache componenti per velocizzare caricamenti ripetuti

def api_componenti():
    global _comp_cache
    p = request.args
    ha_filtri = any(p.get(k) for k in ["q","es_min","es_max","marca","modello"])
    limit  = min(int(p.get("limit",  500)), 2000)
    offset = max(int(p.get("offset", 0)),   0)

    sql = "SELECT * FROM v_giacenza WHERE 1=1"
    params = []
    if p.get("q"):
        t = f"%{p['q']}%"
        search_cols = ["cmp", "articolo", "marca"]
        sql += " AND (" + " OR ".join(f"{c} LIKE ?" for c in search_cols) + ")"
        params += [t] * len(search_cols)
    if p.get("es_min","").isdigit():
        sql += " AND esistenza >= ?"; params.append(int(p["es_min"]))
    if p.get("es_max","").isdigit():
        sql += " AND esistenza <= ?"; params.append(int(p["es_max"]))
    if p.get("sotto_scorta") == "1":
        sql += " AND esistenza = 0"
    if p.get("marca"):
        sql += " AND marca LIKE ?"; params.append(f"%{p['marca']}%")
    if p.get("modello"):
        sql += " AND modello LIKE ?"; params.append(f"%{p['modello']}%")
    sql += " ORDER BY articolo"

    try:
        # Cache full-dataset senza filtri: serve tutte le pagine dalla memoria
        if not ha_filtri and (time.time() - _comp_cache["ts"]) < _COMP_CACHE_TTL and _comp_cache["data"] is not None:
            cached = _comp_cache["data"]
            return _fast_json({"rows": cached[offset:offset+limit], "total": _comp_cache["total"],
                               "offset": offset, "limit": limit, "cached": True})

        conn = db.get_connection()
        try:
            if not ha_filtri:
                # Prima richiesta senza filtri: carica TUTTO in cache (query v_giacenza senza LIMIT)
                all_rows = [dict(r) for r in conn.execute(sql).fetchall()]
                _comp_cache = {"ts": time.time(), "data": all_rows, "total": len(all_rows)}
                return _fast_json({"rows": all_rows[offset:offset+limit], "total": len(all_rows),
                                   "offset": offset, "limit": limit})
            else:
                # Con filtri: COUNT + SELECT limitato
                where_extra = sql[sql.find("WHERE 1=1") + 9:]
                total = conn.execute(
                    "SELECT COUNT(*) AS n FROM v_giacenza WHERE 1=1" + where_extra, params
                ).fetchone()["n"]
                rows = [dict(r) for r in conn.execute(
                    sql + f" LIMIT {limit} OFFSET {offset}", params
                ).fetchall()]
                return _fast_json({"rows": rows, "total": total, "offset": offset, "limit": limit})
        finally:
            conn.close()
    except Exception as e:
        log.warning(f"api_componenti errore: {e}")
        _auto_fix_view()
        try:
            rows = db.fetchall(sql + f" LIMIT {limit} OFFSET {offset}", params)
            return _fast_json({"rows": rows, "total": limit, "offset": offset, "limit": limit})
        except Exception as e2:
            return jsonify({"ok": False, "msg": str(e2)}), 500

@app.route("/api/magazzino/componenti", methods=["POST"])
@require_login
def api_crea_componente():
    dati = request.json or {}
    if not dati.get("codice"):
        return jsonify({"ok": False, "msg": "Codice obbligatorio"}), 400
    if not dati.get("nome") and not dati.get("descrizione"):
        return jsonify({"ok": False, "msg": "Descrizione obbligatoria"}), 400
    if db.fetchone("SELECT id FROM componenti WHERE codice=? AND eliminato=0", (dati["codice"],)):
        return jsonify({"ok": False, "msg": f"Codice '{dati['codice']}' già esistente"}), 400

    u = cu()
    # Mappa campi form → colonne DB (schema Danea 18 col)
    FORM_TO_DB = {
        "codice": "codice", "nome": "descrizione", "descrizione": "descrizione",
        "marca": "produttore", "produttore": "produttore",
        "modello": "modello", "udm": "udm", "cod_udm": "udm",
        "anno": "anno", "anno_da": "anno",
        "cod_prod_forn": "cod_prod_forn",
        "carburante": "alimentazione", "alimentazione": "alimentazione",
        "colore": "colore", "note": "note", "nota": "note",
        "cilindrata": "cilindrata", "ubicazione": "ubicazione",
        "cod_barre": "cod_barre", "extra3": "extra3",
        "cod_fornitore": "cod_fornitore", "fornitore": "fornitore",
        "immagine": "immagine", "immagine_path": "immagine",
    }
    CAMPI_INT   = set()
    CAMPI_FLOAT = set()

    try:
        campi_validi = {}
        for form_key, db_col in FORM_TO_DB.items():
            v = dati.get(form_key)
            if v is None or str(v).strip() == "":
                continue
            if db_col in CAMPI_INT:
                try: v = int(float(v))
                except: continue
            elif db_col in CAMPI_FLOAT:
                try: v = float(str(v).replace(",", "."))
                except: continue
            campi_validi[db_col] = v

        with db._write_lock:
            conn = db.get_connection()
            try:
                cols_str = ", ".join(campi_validi.keys())
                placeholders = ", ".join(["?"] * len(campi_validi))
                cur = conn.execute(
                    f"INSERT INTO componenti ({cols_str}) VALUES ({placeholders})",
                    list(campi_validi.values()))
                comp_id = cur.lastrowid
                conn.execute("""INSERT INTO log_operazioni
                    (utente_id,username,modulo,azione,tabella,record_id,dati_nuovi)
                    VALUES(?,?,?,?,?,?,?)""",
                    (u.get("id"),u.get("username"),"MAGAZZINO","CREA","componenti",comp_id,str(campi_validi)))
                conn.commit()
                conn.close()
                _comp_cache["ts"] = 0  # invalida cache
                return jsonify({"ok": True, "msg": "Componente creato", "id": comp_id})
            except Exception as e:
                conn.rollback()
                return jsonify({"ok": False, "msg": str(e)}), 500
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/magazzino/componenti/<int:cid>", methods=["PUT"])
@require_login
def api_modifica_componente(cid):
    dati = request.json or {}
    u = cu()
    prec = db.fetchone("SELECT * FROM componenti WHERE id=? AND eliminato=0", (cid,))
    if not prec:
        return jsonify({"ok": False, "msg": "Componente non trovato"}), 404

    def v(key, default=None):
        return dati[key] if key in dati else (prec.get(key) if default is None else default)

    try:
        db_write([
            ("""UPDATE componenti SET
                descrizione=?, produttore=?, modello=?, udm=?, anno=?,
                cod_prod_forn=?, alimentazione=?, colore=?, note=?, cilindrata=?,
                ubicazione=?, cod_barre=?, extra3=?,
                cod_fornitore=?, fornitore=?, immagine=?,
                aggiornato_il=datetime('now')
             WHERE id=? AND eliminato=0""",
             (v("descrizione") or v("nome"),
              v("produttore") or v("marca"),
              v("modello"), v("udm", "pz"),
              v("anno") or v("anno_da"),
              v("cod_prod_forn"),
              v("alimentazione") or v("carburante"),
              v("colore"), v("note") or v("nota"), v("cilindrata"),
              v("ubicazione"), v("cod_barre"), v("extra3"),
              v("cod_fornitore"), v("fornitore"),
              v("immagine") or v("immagine_path"),
              cid)),
            ("""INSERT INTO log_operazioni
                (utente_id,username,modulo,azione,tabella,record_id)
                VALUES(?,?,?,?,?,?)""",
             (u.get("id"),u.get("username"),"MAGAZZINO","MODIFICA","componenti",cid))
        ])
        _comp_cache["ts"] = 0  # invalida cache
        return jsonify({"ok": True, "msg": "Componente aggiornato"})
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Errore: {e}"}), 500

@app.route("/api/magazzino/componenti/<int:cid>", methods=["DELETE"])
@require_login
def api_elimina_componente(cid):
    u = cu()
    try:
        db_write([
            ("UPDATE componenti SET eliminato=1, aggiornato_il=datetime('now') WHERE id=?", (cid,)),
            ("""INSERT INTO log_operazioni
                (utente_id,username,modulo,azione,tabella,record_id)
                VALUES(?,?,?,?,?,?)""",
             (u.get("id"),u.get("username"),"MAGAZZINO","ELIMINA","componenti",cid))
        ])
        _comp_cache["ts"] = 0  # invalida cache
        return jsonify({"ok": True, "msg": "Componente eliminato"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/magazzino/prossimo-codice")
@require_login
def api_prossimo_codice():
    import re
    # Legge TUTTI i codici e trova il massimo numerico
    rows = db.fetchall("SELECT codice FROM componenti WHERE eliminato=0")
    max_num = 0
    prefisso = ""
    lunghezza = 0

    for row in rows:
        cod = row.get("codice","") or ""
        m = re.search(r'(\d+)$', cod)
        if m:
            n = int(m.group(1))
            if n > max_num:
                max_num   = n
                prefisso  = cod[:m.start()]
                lunghezza = len(m.group(1))

    if max_num == 0:
        return jsonify({"codice": ""})

    nuovo_num = max_num + 1
    # Mantieni la stessa lunghezza minima del numero (es. 18725 → 18726)
    nuovo = prefisso + str(nuovo_num).zfill(max(lunghezza, len(str(nuovo_num))))
    return jsonify({"codice": nuovo, "ultimo": f"{prefisso}{max_num}"})

# ── Info movimenti per pannello laterale ─────────────────────────────────────
@app.route("/api/magazzino/info-movimenti/<int:cid>")
@require_login
def api_info_movimenti(cid):
    rows = db.fetchall("""
        SELECT tipo, COUNT(*) as cnt, MAX(creato_il) as ultima
        FROM movimenti_magazzino WHERE componente_id=?
        GROUP BY tipo
    """, (cid,))
    r = {"carichi": 0, "scarichi": 0,
         "ultima_data_carico": None, "ultima_data_scarico": None}
    for row in rows:
        if row["tipo"] == "carico":
            r["carichi"] = row["cnt"]; r["ultima_data_carico"] = row["ultima"]
        elif row["tipo"] == "scarico":
            r["scarichi"] = row["cnt"]; r["ultima_data_scarico"] = row["ultima"]
    return jsonify(r)

# ══════════════════════════════════════════════════════════════════════════════
# MOVIMENTI
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/magazzino/movimento", methods=["POST"])
@require_login
def api_movimento():
    d    = request.json or {}
    cid  = d.get("componente_id")
    tipo = d.get("tipo")
    qty  = int(d.get("quantita", 0))
    rif  = d.get("riferimento")
    note = d.get("note")
    u    = cu()

    if not cid or not tipo or qty <= 0:
        return jsonify({"ok": False, "msg": "Dati mancanti"}), 400

    comp = db.fetchone("SELECT esistenza FROM v_giacenza WHERE componente_id=?", (cid,))
    if not comp:
        return jsonify({"ok": False, "msg": "Componente non trovato"}), 404

    gia_prima = comp["esistenza"] or 0
    if tipo == "scarico" and gia_prima < qty:
        return jsonify({"ok": False, "msg": f"Giacenza insufficiente (disponibile: {gia_prima})"}), 400

    gia_dopo = gia_prima + qty if tipo == "carico" else gia_prima - qty

    try:
        with db._write_lock:
            conn = db.get_connection()
            try:
                conn.execute("PRAGMA foreign_keys=OFF")  # disabilita temporaneamente FK
                # Controlla colonne disponibili in movimenti_magazzino
                mov_cols = [r[1] for r in conn.execute("PRAGMA table_info(movimenti_magazzino)").fetchall()]
                # Costruisci insert dinamico
                m_fields = ["componente_id","tipo","quantita","riferimento","note","utente_id"]
                m_vals   = [cid, tipo, qty, rif, note, u.get("id")]
                for col_name, val in [("quantita_prima", gia_prima),("esistenza_prima", gia_prima),
                                      ("quantita_dopo",  gia_dopo), ("esistenza_dopo",  gia_dopo)]:
                    if col_name in mov_cols and col_name not in m_fields:
                        m_fields.append(col_name); m_vals.append(val)
                        break  # usa solo il primo che trova
                for col_name, val in [("quantita_dopo", gia_dopo),("esistenza_dopo", gia_dopo)]:
                    if col_name in mov_cols and col_name not in m_fields:
                        m_fields.append(col_name); m_vals.append(val)
                        break
                conn.execute(
                    "INSERT INTO movimenti_magazzino ("+",".join(m_fields)+") VALUES ("+",".join(["?"]*len(m_fields))+")",
                    m_vals)
                conn.execute("""
                    UPDATE componenti SET giacenza = COALESCE((
                        SELECT SUM(CASE
                            WHEN tipo IN ('carico','inventario') THEN  quantita
                            WHEN tipo = 'scarico'               THEN -quantita
                            WHEN tipo = 'rettifica'             THEN  quantita
                            ELSE 0 END)
                        FROM movimenti_magazzino WHERE componente_id = ?
                    ), 0) WHERE id = ?""", (cid, cid))
                conn.execute("UPDATE magazzino SET aggiornato_il=datetime('now') WHERE componente_id=?", (cid,))
                try:
                    conn.execute(
                        "INSERT INTO log_operazioni (utente_id,username,modulo,azione,tabella,record_id,dati_precedenti,dati_nuovi) VALUES(?,?,?,?,?,?,?,?)",
                        (u.get("id"),u.get("username"),"MAGAZZINO",tipo.upper(),"movimenti_magazzino",cid,str({"g":gia_prima}),str({"g":gia_dopo})))
                except Exception:
                    pass  # log non bloccante
                conn.commit()
            finally:
                conn.close()
        _comp_cache["ts"] = 0  # invalida cache (giacenza cambiata)
        # Recupera id movimento appena inserito
        mov_id = None
        try:
            last = db.fetchone("SELECT id FROM movimenti_magazzino WHERE componente_id=? ORDER BY id DESC LIMIT 1", (cid,))
            if last: mov_id = last["id"]
        except: pass
        return jsonify({"ok": True, "msg": f"{tipo.capitalize()} di {qty} pz completato",
                        "giacenza": gia_dopo, "movimento_id": mov_id})
    except Exception as e:
        log.error(f"api_movimento: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/magazzino/movimenti")
@require_login
def api_movimenti():
    p = request.args
    sql = """SELECT
             m.*,
             c.codice          AS cmp,
             c.descrizione     AS articolo,
             c.produttore,
             c.modello,
             c.colore,
             c.cilindrata,
             c.alimentazione   AS carburante,
             c.ubicazione,
             c.extra3,
             u.username
             FROM movimenti_magazzino m
             LEFT JOIN componenti c ON c.id = m.componente_id
             LEFT JOIN utenti     u ON u.id = m.utente_id
             WHERE 1=1"""
    params = []
    if p.get("componente_id"):
        sql += " AND m.componente_id=?"; params.append(int(p["componente_id"]))
    if p.get("tipo"):
        sql += " AND m.tipo=?"; params.append(p["tipo"])
    if p.get("anno"):
        sql += " AND strftime('%Y',m.creato_il)=?"; params.append(p["anno"])
    if p.get("mese"):
        sql += " AND strftime('%m',m.creato_il)=?"; params.append(p["mese"].zfill(2))
    sql += " ORDER BY m.creato_il DESC"
    # Nessun LIMIT
    return jsonify(db.fetchall(sql, params))

@app.route("/api/magazzino/movimenti/albero")
@require_login
def api_movimenti_albero():
    """Struttura anni → mesi disponibili nello storico."""
    rows = db.fetchall("""
        SELECT DISTINCT strftime('%Y',creato_il) AS anno,
                        strftime('%m',creato_il) AS mese
        FROM movimenti_magazzino
        ORDER BY anno DESC, mese DESC
    """)
    tree = {}
    for r in rows:
        a, m = r["anno"], r["mese"]
        tree.setdefault(a, [])
        if m not in tree[a]: tree[a].append(m)
    return jsonify(tree)

# ══════════════════════════════════════════════════════════════════════════════
# STATS / INVENTARIO / LISTA ACQUISTI
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/magazzino/movimenti/<int:mid>", methods=["DELETE"])
@require_login
def api_elimina_movimento(mid):
    """Elimina fisicamente un movimento (usato da undo)."""
    u = cu()
    try:
        mov = db.fetchone("SELECT * FROM movimenti_magazzino WHERE id=?", (mid,))
        if not mov:
            return jsonify({"ok": False, "msg": "Movimento non trovato"}), 404
        with db._write_lock:
            conn = db.get_connection()
            try:
                conn.execute("DELETE FROM movimenti_magazzino WHERE id=?", (mid,))
                conn.execute("""
                    UPDATE componenti SET giacenza = COALESCE((
                        SELECT SUM(CASE
                            WHEN tipo IN ('carico','inventario') THEN  quantita
                            WHEN tipo = 'scarico'               THEN -quantita
                            WHEN tipo = 'rettifica'             THEN  quantita
                            ELSE 0 END)
                        FROM movimenti_magazzino WHERE componente_id = ?
                    ), 0) WHERE id = ?""", (mov["componente_id"], mov["componente_id"]))
                conn.execute("""INSERT INTO log_operazioni
                    (utente_id,username,modulo,azione,tabella,record_id)
                    VALUES(?,?,?,?,?,?)""",
                    (u.get("id"),u.get("username"),"MAGAZZINO","UNDO_DELETE","movimenti_magazzino",mid))
                conn.commit()
            finally:
                conn.close()
        return jsonify({"ok": True, "msg": "Movimento annullato", "movimento": dict(mov)})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

def _get_lan_ip():
    """Restituisce l'IP LAN della macchina (non localhost)."""
    import socket as _sock
    try:
        s = _sock.socket(_sock.AF_INET, _sock.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

@app.route("/api/server-info")
def api_server_info():
    port = int(os.environ.get("PORT", 5000))
    return jsonify({"lan_ip": _get_lan_ip(), "port": port})

@app.route("/api/magazzino/stats")
@require_login
def api_stats():
    t    = db.fetchone("SELECT COUNT(*) AS n FROM componenti WHERE eliminato=0")
    u    = db.fetchone("SELECT creato_il FROM movimenti_magazzino ORDER BY id DESC LIMIT 1")
    pz   = db.fetchone("SELECT COALESCE(SUM(giacenza),0) AS n FROM componenti WHERE eliminato=0")
    disp = db.fetchone("SELECT COUNT(*) AS n FROM componenti WHERE eliminato=0 AND giacenza > 0")
    foto = db.fetchone("SELECT COUNT(*) AS n FROM componenti WHERE immagine IS NOT NULL AND immagine!='' AND eliminato=0")
    return jsonify({"totale_componenti": t["n"] if t else 0,
                    "disponibili": disp["n"] if disp else 0,
                    "ultimo_movimento": u["creato_il"] if u else "—",
                    "pezzi_totali": pz["n"] if pz else 0,
                    "con_foto": foto["n"] if foto else 0})

@app.route("/api/backup", methods=["POST"])
@require_login
def api_backup():
    import shutil, datetime
    try:
        src  = cfg.get("db_path")
        bdir = ROOT / "backup"
        bdir.mkdir(exist_ok=True)
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dst  = bdir / f"perilcar_{ts}.db"
        shutil.copy2(src, dst)
        backups = sorted(bdir.glob("perilcar_*.db"))
        for old in backups[:-30]:
            old.unlink()
        u = cu()
        db_write([("INSERT INTO log_operazioni(utente_id,username,modulo,azione) VALUES(?,?,?,?)",
                   (u.get("id"), u.get("username"), "SISTEMA", "BACKUP"))])
        return jsonify({"ok": True, "msg": f"Backup salvato: {dst.name}", "file": dst.name})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/magazzino/inventario")
@require_login
def api_inventario():
    return jsonify(db.fetchall("""
        SELECT cmp, articolo, esistenza,
               marca, modello, colore,
               cilindrata, carburante,
               anno_da, nota, ubicazione
        FROM v_giacenza ORDER BY articolo
    """))

@app.route("/api/magazzino/lista-acquisti")
@require_login
def api_lista_acquisti():
    p = request.args
    marca   = p.get("marca","").strip()
    modello = p.get("modello","").strip()
    try: soglia = int(p.get("soglia", 1))
    except: soglia = 1

    sql = "SELECT * FROM v_giacenza WHERE esistenza < ?"
    params = [soglia]
    if marca:
        sql += " AND marca LIKE ?"; params.append(f"%{marca}%")
    if modello:
        sql += " AND modello LIKE ?"; params.append(f"%{modello}%")
    sql += " ORDER BY articolo"
    return jsonify(db.fetchall(sql, params))

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD IMMAGINI
# ══════════════════════════════════════════════════════════════════════════════

ALLOWED = {"png","jpg","jpeg","gif","webp","pdf","xlsx","xls","docx","zip"}

@app.route("/api/magazzino/upload-file/<int:cid>", methods=["POST"])
@require_login
def api_upload_file(cid):
    """Upload immagini o file allegati a un componente (anche multipli)."""
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f   = request.files["file"]
    ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else ""
    if ext not in ALLOWED:
        return jsonify({"ok": False, "msg": f"Tipo .{ext} non permesso"}), 400
    import uuid
    fname = f"comp_{cid}_{uuid.uuid4().hex[:8]}.{ext}"
    f.save(UPLOAD_FOLDER / fname)
    url = f"/static/uploads/{fname}"
    # Imposta immagine principale se è un'immagine e non ce n'è una
    imgs_ext = {"png","jpg","jpeg","gif","webp"}
    comp = db.fetchone("SELECT immagine FROM componenti WHERE id=?", (cid,))
    new_img = comp["immagine"] if comp else None
    if ext in imgs_ext and not new_img:
        new_img = url
    try:
        db_write([("""UPDATE componenti
                      SET immagine=?, aggiornato_il=datetime('now')
                      WHERE id=?""", (new_img, cid))])
        return jsonify({"ok": True, "url": url, "files": [url]})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/magazzino/delete-file/<int:cid>", methods=["POST"])
@require_login
def api_delete_file(cid):
    """Rimuove immagine da un componente."""
    url_da_rimuovere = (request.json or {}).get("url","")
    comp = db.fetchone("SELECT immagine FROM componenti WHERE id=?", (cid,))
    new_img = None if (comp and comp["immagine"] == url_da_rimuovere) else (comp["immagine"] if comp else None)
    try:
        db_write([("""UPDATE componenti SET immagine=?, aggiornato_il=datetime('now') WHERE id=?""",
                   (new_img, cid))])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

# ══════════════════════════════════════════════════════════════════════════════
# IMPORT / EXPORT EXCEL — formato Danea + generico
# ══════════════════════════════════════════════════════════════════════════════

# ── Rilevamento automatico colonne Excel ─────────────────────────────────────
# Ogni entry: (campo_db, [alias_lista], tipo, label_ui)
# tipo: "chiave" | "richiesto" | "movimento" | "calcolo" | "prezzo" | None
_CAMPO_ALIASES = [
    ("codice",        ["cod.", "cod", "codice", "sku", "cmp", "codart", "cod articolo",
                       "codice articolo", "cod. articolo", "riferimento", "rif", "rif.",
                       "code", "art.", "art", "numero articolo", "n. articolo",
                       "cod.art.", "cod. art.", "codice art", "codice ricambio"],
                      "chiave",    "Codice articolo"),
    ("nome",          ["nome", "descrizione", "articolo", "desc", "descrizione articolo",
                       "nome articolo", "denominazione", "item", "prodotto",
                       "articolo desc", "descrizione completa", "descrizione prodotto"],
                      "richiesto", "Nome / Descrizione"),
    ("esistenza",     ["quantita", "quantità", "qta", "q.ta", "q.tà", "q.tà giacenza",
                       "giacenza", "esistenza", "es", "qty", "disponibile", "stock",
                       "pezzi", "qt", "quantita giacenza", "q.ta giacenza",
                       "giacenza attuale", "disponibilita", "disponibilità",
                       "magazzino", "scorta attuale"],
                      "movimento", "Quantità (giacenza iniziale)"),
    ("scorta_minima", ["scorta min", "scorta minima", "scorta min.", "scorta min. (pz)",
                       "min", "minimo", "qtà minima", "giacenza minima", "stock minimo",
                       "qt min", "qt. min", "quantita minima"],
                      "calcolo",   "Scorta minima"),
    ("categoria",     ["categoria", "cat", "category", "gruppo", "reparto", "famiglia"],
                      None,        "Categoria"),
    ("sottocategoria",["sottocategoria", "sottocat", "sotto categoria", "subcategoria"],
                      None,        "Sottocategoria"),
    ("tipologia",     ["tipologia", "tipo", "type", "kind"],
                      None,        "Tipologia"),
    ("marca",         ["marca", "brand", "produttore", "fabbricante", "manufacturer"],
                      None,        "Marca / Produttore"),
    ("modello",       ["modello", "model", "mod.", "mod"],
                      None,        "Modello"),
    ("nota",          ["nota", "note", "osservazioni", "commento", "commenti", "info",
                       "annotazioni", "descrizione lunga", "note aggiuntive"],
                      None,        "Note"),
    ("ubicazione",    ["ubicazione", "posizione", "pos", "location", "scaffale",
                       "zona", "posizione magazzino", "ubicaz."],
                      None,        "Ubicazione"),
    ("colore",        ["colore", "color", "col."],
                      None,        "Colore"),
    ("cod_barre",     ["cod barre", "cod. a barre", "barcode", "ean", "codice barre",
                       "cod. barre", "upc", "cod.barre", "ean13"],
                      None,        "Cod. a barre"),
    ("listino1",      ["listino 1", "listino1", "prezzo", "prezzo vendita", "p.v.",
                       "prezzo 1", "pr vendita", "vendita", "prezzo al pubblico",
                       "prezzo cliente", "pr. vendita"],
                      "prezzo",    "Listino 1 (prezzo vendita)"),
    ("listino2",      ["listino 2", "listino2", "prezzo 2", "p.v. 2", "listino b"],
                      "prezzo",    "Listino 2"),
    ("listino3",      ["listino 3", "listino3", "prezzo 3", "p.v. 3", "listino c"],
                      "prezzo",    "Listino 3"),
    ("prezzo_forn",   ["prezzo forn.", "prezzo fornitore", "costo", "costo acquisto",
                       "p.a.", "prezzo acquisto", "pr acquisto", "costo unitario",
                       "prezzo forn", "costo forn"],
                      "prezzo",    "Prezzo fornitore / Costo"),
    ("cod_fornitore", ["cod fornitore", "cod. fornitore", "codice fornitore", "cod forn"],
                      None,        "Cod. fornitore"),
    ("fornitore",     ["fornitore", "supplier", "vendor", "fornitore principale"],
                      None,        "Fornitore"),
    ("stato_magazzino",["stato magazzino", "stato", "status"],
                      None,        "Stato magazzino"),
    ("anno_da",       ["anno da", "anno dal", "anno inizio", "anno from", "from year"],
                      None,        "Anno da"),
    ("anno_a",        ["anno a", "anno al", "anno fine", "anno to", "to year"],
                      None,        "Anno a"),
    ("cod_udm",       ["cod udm", "cod. udm", "udm", "unità misura", "um",
                       "unita misura", "u.m."],
                      None,        "Unità di misura"),
    ("cod_iva",       ["cod iva", "cod. iva", "iva", "aliquota", "aliquota iva",
                       "% iva"],
                      None,        "Cod. IVA"),
    ("internet",      ["internet", "web", "online", "sito"],
                      None,        "Internet/Web"),
    ("extra1",        ["extra1", "extra 1", "campo extra 1", "campo1"],
                      None,        "Extra 1"),
    ("extra2",        ["extra2", "extra 2", "campo extra 2", "campo2"],
                      None,        "Extra 2"),
    ("extra3",        ["extra3", "extra 3", "campo extra 3", "campo3"],
                      None,        "Extra 3"),
    ("extra4",        ["extra4", "extra 4", "campo extra 4", "campo4"],
                      None,        "Extra 4"),
]

_CAMPO_INFO = {c: {"tipo": t, "label": l, "alias": a}
               for c, a, t, l in _CAMPO_ALIASES}


def _rileva_mapping(headers):
    """Rileva automaticamente a cosa corrisponde ogni colonna Excel.
    Restituisce dict {header_originale: {campo, label, tipo, confidenza}}.
    """
    result = {}
    used_fields = set()

    def _norm(s):
        return " ".join(s.lower().replace(".", " ").replace("'", " ")
                         .replace("_", " ").replace("(", " ").replace(")", " ")
                         .split())

    for h in headers:
        h_norm = _norm(h)
        best_field = None
        best_score = 0

        for campo, aliases, tipo, label in _CAMPO_ALIASES:
            if campo in used_fields:
                continue
            for alias in aliases:
                a_norm = _norm(alias)
                if h_norm == a_norm:
                    score = 100
                elif a_norm in h_norm or h_norm in a_norm:
                    score = 70
                else:
                    continue
                if score > best_score:
                    best_score = score
                    best_field = campo

        if not best_field:
            danea = DANEA_MAP.get(h.lower().strip())
            if danea:
                best_field = danea
                best_score = 80

        if best_field and best_field not in used_fields:
            used_fields.add(best_field)
            info = _CAMPO_INFO.get(best_field, {})
            result[h] = {
                "campo":      best_field,
                "label":      info.get("label", best_field.replace("_", " ").title()),
                "tipo":       info.get("tipo"),
                "confidenza": "alta" if best_score >= 100 else "media",
            }
        else:
            result[h] = {
                "campo":      None,
                "label":      "— Ignora —",
                "tipo":       None,
                "confidenza": "nessuna",
            }

    return result


@app.route("/api/magazzino/analizza-excel", methods=["POST"])
@require_login
def api_analizza_excel():
    """Legge le colonne dell'Excel e suggerisce il mapping senza importare."""
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "msg": "Solo .xlsx o .xls"}), 400
    import openpyxl
    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h or "").strip() for h in header_row if h is not None and str(h).strip()]
        totale = (ws.max_row or 1) - 1
        mapping = _rileva_mapping(headers)
        campi_disponibili = (
            [{"campo": c, "label": info["label"], "tipo": info.get("tipo")}
             for c, info in _CAMPO_INFO.items()]
            + [{"campo": None, "label": "— Ignora —", "tipo": None}]
        )
        return jsonify({"ok": True, "headers": headers, "mapping": mapping,
                        "campi_disponibili": campi_disponibili, "totale": totale})
    except Exception as e:
        log.warning(f"analizza-excel: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500


def _processa_import_mappato(tmp_path, mapping, utente_id, utente_username):
    """Import Excel con mapping colonne confermato dall'utente."""
    global _import_stato
    import openpyxl, os as _os, datetime as _dt

    _import_stato.update({"running": True, "importati": 0, "aggiornati": 0,
                           "saltati": 0, "totale": 0, "processed": 0,
                           "msg": "Avvio importazione...", "ok": None})
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h or "").strip() for h in header_row]

        # campo_db → indice colonna Excel
        campo_idx = {}
        for i, h in enumerate(headers):
            campo = mapping.get(h)
            if campo and campo not in campo_idx:
                campo_idx[campo] = i

        def get_val(row, campo, default=None):
            idx = campo_idx.get(campo)
            if idx is None or idx >= len(row): return default
            v = row[idx]
            if v is None: return default
            if isinstance(v, (_dt.datetime, _dt.date)): return str(v.year)
            s = str(v).strip()
            return s if s not in ("None", "nan", "") else default

        def to_int(row, campo):
            try: v = get_val(row, campo, "0"); return int(float(v)) if v else 0
            except: return 0

        def to_float(row, campo):
            try:
                v = get_val(row, campo)
                return float(str(v).replace(",", ".")) if v else None
            except: return None

        col_codice = _DB_COLS.get("cmp", "codice")
        col_nome   = _DB_COLS.get("articolo", "nome")
        col_nota   = _DB_COLS.get("nota", "nota")
        col_scorta = _DB_COLS.get("scorta", "scorta_minima")
        col_elim   = _DB_COLS.get("eliminato", "eliminato")
        tab_mov    = _DB_COLS.get("tabella_movimenti", "movimenti_magazzino")

        CAMPI_FLOAT = {"listino1", "listino2", "listino3", "prezzo_forn"}
        CAMPI_INT   = {"scorta_min", "ord_multipli", "gg_ordine"}

        importati = aggiornati = saltati = row_n = 0
        BATCH = 500

        with db._write_lock:
            conn = db.get_connection()
            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_n += 1
                    codice    = get_val(row, "codice")
                    nome      = get_val(row, "nome")
                    esistenza = to_int(row, "esistenza")

                    if not codice and not nome:
                        saltati += 1; continue
                    if not codice: codice = f"IMP-{row_n:05d}"
                    if not nome:   nome   = codice

                    campi = {col_nome: nome}

                    for campo in campo_idx:
                        if campo in ("codice", "nome", "esistenza"):
                            continue
                        if campo in CAMPI_FLOAT:
                            v = to_float(row, campo)
                        elif campo in CAMPI_INT:
                            v = to_int(row, campo)
                        else:
                            v = get_val(row, campo)

                        if v is None or v == "":
                            continue

                        db_col = col_nota   if campo == "nota"      else \
                                 col_scorta if campo == "scorta_min" else campo
                        campi[db_col] = v

                    existing = conn.execute(
                        f"SELECT id FROM componenti WHERE {col_codice}=? AND {col_elim}=0",
                        (codice,)
                    ).fetchone()

                    if existing:
                        comp_id = existing[0]
                        if campi:
                            sets = ", ".join(f"{k}=?" for k in campi)
                            conn.execute(
                                f"UPDATE componenti SET {sets}, aggiornato_il=datetime('now') WHERE id=?",
                                list(campi.values()) + [comp_id]
                            )
                        aggiornati += 1
                    else:
                        campi[col_codice] = codice
                        campi[col_elim]   = 0
                        cols_str = ", ".join(campi.keys())
                        phs      = ", ".join(["?"] * len(campi))
                        cur = conn.execute(
                            f"INSERT INTO componenti ({cols_str}) VALUES ({phs})",
                            list(campi.values())
                        )
                        comp_id = cur.lastrowid
                        importati += 1

                    if esistenza > 0:
                        already = conn.execute(
                            f"SELECT id FROM {tab_mov} WHERE componente_id=? AND riferimento='Import Excel' LIMIT 1",
                            (comp_id,)
                        ).fetchone()
                        if not already:
                            try:
                                conn.execute(
                                    f"INSERT INTO {tab_mov} (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento,username) VALUES(?,?,?,?,?,?,?)",
                                    (comp_id, "carico", esistenza, 0, esistenza, "Import Excel", utente_username)
                                )
                            except Exception:
                                conn.execute(
                                    f"INSERT INTO {tab_mov} (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento) VALUES(?,?,?,?,?,?)",
                                    (comp_id, "carico", esistenza, 0, esistenza, "Import Excel")
                                )
                            conn.execute(
                                "UPDATE componenti SET giacenza=giacenza+? WHERE id=?",
                                (esistenza, comp_id))

                    if row_n % BATCH == 0:
                        conn.commit()
                        _import_stato.update({
                            "importati": importati, "aggiornati": aggiornati,
                            "saltati": saltati, "processed": row_n,
                            "msg": f"⏳ {importati} nuovi, {aggiornati} aggiornati, riga {row_n}..."
                        })
                        socketio.emit("import_progress", {
                            "importati": importati, "aggiornati": aggiornati,
                            "processed": row_n, "saltati": saltati
                        })

                conn.commit()
            except Exception as _e:
                log.error(f"Import mappato errore riga {row_n}: {_e}")
                try: conn.rollback()
                except: pass
                raise
            finally:
                conn.close()

        try: _os.remove(tmp_path)
        except: pass

        msg = f"✅ Import completato: {importati} nuovi, {aggiornati} aggiornati, {saltati} saltati su {row_n} righe"
        log.info(msg)
        _comp_cache["ts"] = 0  # invalida cache dopo import
        _import_stato.update({"running": False, "importati": importati,
                               "aggiornati": aggiornati, "saltati": saltati,
                               "processed": row_n, "msg": msg, "ok": True})
        socketio.emit("import_done", {"ok": True, "msg": msg,
                                       "importati": importati, "aggiornati": aggiornati,
                                       "saltati": saltati})
    except Exception as e:
        msg = f"❌ Errore import: {e}"
        log.error(msg)
        _import_stato.update({"running": False, "msg": msg, "ok": False})
        socketio.emit("import_done", {"ok": False, "msg": msg})


@app.route("/api/magazzino/import-excel-mappato", methods=["POST"])
@require_login
def api_import_excel_mappato():
    """Avvia import Excel con mapping confermato dall'utente."""
    global _import_stato
    if _import_stato.get("running"):
        return jsonify({"ok": False, "msg": "Import già in corso"}), 409
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "msg": "Solo .xlsx o .xls"}), 400
    import json as _json, os as _os
    try:
        mapping = _json.loads(request.form.get("mapping", "{}"))
    except Exception:
        return jsonify({"ok": False, "msg": "Mapping JSON non valido"}), 400
    u = cu()
    tmp_path = str(ROOT / "logs" / "_import_mappato_tmp.xlsx")
    f.save(tmp_path)
    log.info(f"Import mappato: file {_os.path.getsize(tmp_path)//1024}KB, "
             f"{len(mapping)} colonne mappate")
    t = threading.Thread(target=_processa_import_mappato,
                         args=(tmp_path, mapping, u.get("id"), u.get("username")),
                         daemon=True)
    t.start()
    return jsonify({"ok": True, "msg": "Import avviato"})


# Mappa colonne Danea → campi DB (case-insensitive, schema 18 col)
DANEA_MAP = {
    "cod.":             "codice",
    "descrizione":      "descrizione",
    "produttore":       "produttore",
    "modello":          "modello",
    "q.tà giacenza":    "esistenza",
    "pz":               "udm",
    "anno":             "anno",
    "cod. prod. forn.": "cod_prod_forn",
    "alimentazione":    "alimentazione",
    "colore":           "colore",
    "note":             "note",
    "cilindrata":       "cilindrata",
    "ubicazione":       "ubicazione",
    "cod. a barre":     "cod_barre",
    "extra 3":          "extra3",
    "cod. fornitore":   "cod_fornitore",
    "fornitore":        "fornitore",
    "immagine":         "immagine",
    # alias generici
    "cmp":              "codice",
    "articolo":         "descrizione",
    "nome":             "descrizione",
    "es":               "esistenza",
    "giacenza":         "esistenza",
    "anno da":          "anno",
    "marca":            "produttore",
    "nota":             "note",
    "carburante":       "alimentazione",
    "cod. udm":         "udm",
    "udm":              "udm",
}

@app.route("/api/magazzino/export-excel")
@require_login
def api_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    import tempfile, os

    rows = db.fetchall("SELECT * FROM v_giacenza ORDER BY articolo")
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Magazzino"

    # Stili
    hdr_fill = PatternFill("solid", fgColor="E94C00")
    alt_fill = PatternFill("solid", fgColor="FFF3EE")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    cell_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    cols = [
        ("Cod.",             "cmp",            10),
        ("Descrizione",      "articolo",       36),
        ("Produttore",       "marca",          16),
        ("Modello",          "modello",        16),
        ("Q.tà giacenza",    "esistenza",      12),
        ("Cod. Udm",         "cod_udm",        10),
        ("Anno",             "anno_da",        10),
        ("Cod. prod. forn.", "cod_prod_forn",  16),
        ("Alimentazione",    "carburante",     14),
        ("Colore",           "colore",         12),
        ("Note",             "nota",           30),
        ("Cilindrata",       "cilindrata",     12),
        ("Ubicazione",       "ubicazione",     18),
        ("Cod. a barre",     "cod_barre",      14),
        ("Extra 3",          "extra3",         12),
        ("Cod. fornitore",   "cod_fornitore",  14),
        ("Fornitore",        "fornitore",      16),
        ("Foto",             "immagine_path",  20),
    ]

    NUM_KEYS = {"esistenza", "anno_da"}

    # Intestazioni
    for c, (label, _, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = center
        cell.border     = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Dati
    foto_col_idx = next((c+1 for c,(l,k,w) in enumerate(cols) if k=="immagine_path"), None)
    tmp_files = []

    for ri, r in enumerate(rows, 2):
        is_alt = (ri % 2 == 0)
        ws.row_dimensions[ri].height = 15  # altezza normale

        for ci, (_, key, _) in enumerate(cols, 1):
            if key == "immagine_path":
                # Scrivi URL immagine principale
                paths = [r.get("immagine_path")] if r.get("immagine_path") else []
                foto_cell = ws.cell(row=ri, column=ci, value="")
                foto_cell.border    = border
                foto_cell.alignment = center
                if paths:
                    # Inserisci immagine
                    img_url = paths[0]
                    if img_url.startswith("/static/"):
                        img_path = ROOT / "web" / img_url[1:]
                    else:
                        img_path = None
                    if img_path and img_path.exists():
                        try:
                            # Ridimensiona a 60x60 con PIL
                            pil = PILImage.open(str(img_path))
                            pil.thumbnail((60, 60), PILImage.LANCZOS)
                            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                            pil.save(tmp.name, "PNG")
                            tmp.close()
                            tmp_files.append(tmp.name)
                            img = XLImage(tmp.name)
                            img.anchor = get_column_letter(ci) + str(ri)
                            ws.add_image(img)
                            ws.row_dimensions[ri].height = 48
                        except Exception as e:
                            log.warning(f"Foto Excel fallita: {e}")
                            cell.value = "📷"
            else:
                val = r.get(key)
                if val is None or val == "":
                    val = 0 if key in NUM_KEYS else ""
                c2 = ws.cell(row=ri, column=ci, value=val)
                c2.font      = cell_font
                c2.border    = border
                c2.alignment = center if key in NUM_KEYS else left
                if is_alt:
                    c2.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Pulisci file temporanei
    for f in tmp_files:
        try: os.unlink(f)
        except: pass

    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="perilcar_magazzino.xlsx")


@app.route("/api/magazzino/import-excel", methods=["POST"])
@require_login
def api_import_excel():
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx",".xls")):
        return jsonify({"ok": False, "msg": "Solo file .xlsx o .xls"}), 400

    import openpyxl
    u = cu()

    try:
        wb   = openpyxl.load_workbook(f, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return jsonify({"ok": False, "msg": "File vuoto o senza dati"}), 400

        # Mappa intestazioni → indice colonna
        header  = [str(h or "").strip().lower() for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            field = DANEA_MAP.get(h, h.replace(" ","_").replace(".","").replace("'",""))
            col_map[field] = i

        def get(row, field, default=None):
            idx = col_map.get(field)
            if idx is None or idx >= len(row): return default
            v = row[idx]
            if v is None: return default
            # datetime → stringa
            import datetime
            if isinstance(v, (datetime.datetime, datetime.date)):
                return str(v.year)
            return str(v).strip() if str(v).strip() not in ("None","nan","") else default

        def toint(row, field):
            try: v = get(row, field, "0"); return int(float(v)) if v else 0
            except: return 0

        def tofloat(row, field):
            try: v = get(row, field, "0"); return float(v) if v else 0.0
            except: return 0.0

        importati = 0; saltati = 0; aggiornati = 0; row_n = 1

        with db._write_lock:
            conn = db.get_connection()
            try:
                for row in rows[1:]:
                    row_n += 1
                    codice = get(row, "codice")
                    nome   = get(row, "nome")
                    if not codice:
                        codice = get(row,"cmp") or get(row,"ref") or get(row,"riferimento")
                    if not nome:
                        nome = get(row,"descrizione") or get(row,"prodotto") or get(row,"articolo")
                    if not codice and not nome:
                        saltati += 1; continue
                    if not codice: codice = "IMP-" + str(row_n).zfill(4)
                    if not nome:   nome   = codice

                    esistenza = toint(row, "esistenza")

                    campi_raw = {
                        "descrizione":   nome,
                        "udm":           get(row, "udm") or get(row, "cod_udm") or "pz",
                        "note":          get(row, "nota") or get(row, "note"),
                        "cod_barre":     get(row, "cod_barre"),
                        "produttore":    get(row, "produttore") or get(row, "marca"),
                        "modello":       get(row, "modello"),
                        "extra3":        get(row, "extra3"),
                        "cod_fornitore": get(row, "cod_fornitore"),
                        "fornitore":     get(row, "fornitore"),
                        "cod_prod_forn": get(row, "cod_prod_forn"),
                        "ubicazione":    get(row, "ubicazione"),
                        "anno":          get(row, "anno") or get(row, "anno_da"),
                        "alimentazione": get(row, "alimentazione") or get(row, "carburante"),
                        "colore":        get(row, "colore"),
                        "cilindrata":    get(row, "cilindrata"),
                    }
                    valid_cols = set(_DB_COLS.get("all_cols", []))
                    campi = {k: v for k, v in campi_raw.items()
                             if v is not None and v != "" and (not valid_cols or k in valid_cols)}

                    existing = conn.execute(
                        "SELECT id FROM componenti WHERE codice=? AND eliminato=0",
                        (codice,)).fetchone()

                    if existing:
                        comp_id = existing[0]
                        sets    = ", ".join(f"{k}=?" for k in campi)
                        vals    = list(campi.values()) + [comp_id]
                        conn.execute(
                            f"UPDATE componenti SET {sets}, aggiornato_il=datetime('now') WHERE id=?",
                            vals)
                        aggiornati += 1
                    else:
                        campi["codice"]   = codice
                        campi["eliminato"] = 0
                        cols_str = ", ".join(campi.keys())
                        placeholders = ", ".join(["?"] * len(campi))
                        cur = conn.execute(
                            f"INSERT INTO componenti ({cols_str}) VALUES ({placeholders})",
                            list(campi.values()))
                        comp_id = cur.lastrowid
                        conn.execute(
                            "INSERT OR IGNORE INTO magazzino(componente_id) VALUES(?)",
                            (comp_id,))
                        importati += 1

                    # Carico iniziale se giacenza > 0
                    if esistenza > 0:
                        already = conn.execute(
                            "SELECT id FROM movimenti_magazzino WHERE componente_id=? AND riferimento='Import Excel' LIMIT 1",
                            (comp_id,)).fetchone()
                        if not already:
                            conn.execute("""INSERT INTO movimenti_magazzino
                                (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento,utente_id)
                                VALUES(?,?,?,?,?,?,?)""",
                                (comp_id,"carico",esistenza,0,esistenza,"Import Excel",u.get("id")))

                conn.commit()
                conn.close()
            except Exception as e:
                conn.rollback()
                log.error(f"Import errore riga {row_n}: {e}")
                return jsonify({"ok": False, "msg": f"Errore riga {row_n}: {e}"}), 500

        tot = importati + aggiornati
        return jsonify({"ok": True,
            "msg": f"✅ Import OK: {importati} nuovi, {aggiornati} aggiornati, {saltati} saltati su {tot} totali",
            "importati": importati, "aggiornati": aggiornati, "saltati": saltati})

    except Exception as e:
        log.error(f"Import lettura file: {e}")
        return jsonify({"ok": False, "msg": f"Errore lettura file: {e}"}), 500


@app.route("/api/magazzino/import-excel-stream", methods=["POST"])
@require_login
def api_import_excel_stream():
    """Import Excel - salva su disco locale poi processa per evitare timeout."""
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx",".xls")):
        return jsonify({"ok": False, "msg": "Solo .xlsx o .xls"}), 400

    import openpyxl, tempfile, os as _os
    u = cu()

    # Salva su disco
    tmp_path = str(ROOT / "db" / "_import_tmp.xlsx")
    f.save(tmp_path)
    log.info(f"File salvato: {tmp_path} ({_os.path.getsize(tmp_path)//1024}KB)")

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h or "").strip().lower() for h in header_row]
        col_map = {}
        for i, h in enumerate(header):
            field = DANEA_MAP.get(h, h.replace(" ","_").replace(".","").replace("'",""))
            col_map[field] = i

        def get(row, field, default=None):
            idx = col_map.get(field)
            if idx is None or idx >= len(row): return default
            v = row[idx]
            if v is None: return default
            import datetime
            if isinstance(v, (datetime.datetime, datetime.date)):
                return str(v.year)
            s = str(v).strip()
            return s if s not in ("None","nan","") else default

        def toint(row, field):
            try: v = get(row, field, "0"); return int(float(v)) if v else 0
            except: return 0

        def tofloat(row, field):
            try: v = get(row, field, "0"); return float(v) if v else 0.0
            except: return 0.0

        importati = 0; aggiornati = 0; saltati = 0; row_n = 1
        BATCH = 200

        with db._write_lock:
            conn = db.get_connection()
            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_n += 1
                    codice = get(row, "codice")
                    nome   = get(row, "nome")
                    if not codice:
                        codice = get(row,"cmp") or get(row,"ref") or get(row,"riferimento")
                    if not nome:
                        nome = get(row,"descrizione") or get(row,"prodotto") or get(row,"articolo")
                    if not codice and not nome:
                        saltati += 1; continue
                    if not codice: codice = "IMP-" + str(row_n).zfill(4)
                    if not nome:   nome   = codice

                    esistenza = toint(row, "esistenza")

                    campi_raw = {
                        "descrizione":   nome,
                        "udm":           get(row,"udm") or get(row,"cod_udm") or "pz",
                        "note":          get(row,"nota") or get(row,"note"),
                        "cod_barre":     get(row,"cod_barre"),
                        "produttore":    get(row,"produttore") or get(row,"marca"),
                        "modello":       get(row,"modello"),
                        "extra3":        get(row,"extra3"),
                        "cod_fornitore": get(row,"cod_fornitore"),
                        "fornitore":     get(row,"fornitore"),
                        "cod_prod_forn": get(row,"cod_prod_forn"),
                        "ubicazione":    get(row,"ubicazione"),
                        "colore":        get(row,"colore"),
                        "anno":          get(row,"anno") or get(row,"anno_da"),
                        "alimentazione": get(row,"alimentazione") or get(row,"carburante"),
                        "cilindrata":    get(row,"cilindrata"),
                    }
                    valid_cols = set(_DB_COLS.get("all_cols", []))
                    campi = {k:v for k,v in campi_raw.items()
                             if v is not None and v != "" and (not valid_cols or k in valid_cols)}

                    col_cmp_r  = _DB_COLS.get("cmp","cmp")
                    col_elim_r = _DB_COLS.get("eliminato","eliminato")
                    col_aggi_r = _DB_COLS.get("aggiornato_il","aggiornato_il")
                    all_cols_r = _DB_COLS.get("all_cols",[])

                    existing = conn.execute(
                        f"SELECT id FROM componenti WHERE {col_cmp_r}=?"
                        + (f" AND {col_elim_r}=0" if col_elim_r else ""),
                        (codice,)
                    ).fetchone()

                    if existing:
                        comp_id = existing[0]
                        if campi:
                            sets = ", ".join(f"{k}=?" for k in campi)
                            conn.execute(
                                f"UPDATE componenti SET {sets}, {col_aggi_r}=datetime('now') WHERE id=?",
                                list(campi.values()) + [comp_id])
                        aggiornati += 1
                    else:
                        campi_ins = dict(campi)
                        campi_ins[col_cmp_r] = codice
                        if col_elim_r and (not all_cols_r or col_elim_r in all_cols_r):
                            campi_ins[col_elim_r] = 0
                        cols_str = ", ".join(campi_ins.keys())
                        phs = ", ".join(["?"] * len(campi_ins))
                        cur = conn.execute(
                            f"INSERT INTO componenti ({cols_str}) VALUES ({phs})",
                            list(campi_ins.values()))
                        comp_id = cur.lastrowid
                        importati += 1

                    if esistenza > 0:
                        already = conn.execute(
                            "SELECT id FROM movimenti_magazzino WHERE componente_id=? AND riferimento='Import Excel' LIMIT 1",
                            (comp_id,)).fetchone()
                        if not already:
                            conn.execute(
                                "INSERT INTO movimenti_magazzino (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento,utente_id) VALUES(?,?,?,?,?,?,?)",
                                (comp_id,"carico",esistenza,0,esistenza,"Import Excel",u.get("id")))

                    if row_n % BATCH == 0:
                        conn.commit()
                        conn.close()
                        log.info(f"Import progress: riga {row_n}, nuovi={importati}, aggiornati={aggiornati}")
                        socketio.emit("import_progress", {
                            "processed": row_n-1,
                            "importati": importati,
                            "aggiornati": aggiornati
                        })

                conn.commit()
                conn.close()
                wb.close()
            except Exception as e:
                conn.rollback()
                log.error(f"Import errore riga {row_n}: {e}")
                return jsonify({"ok": False, "msg": f"Errore riga {row_n}: {e}"}), 500

    finally:
        try: _os.unlink(tmp_path)
        except: pass

    tot = importati + aggiornati
    socketio.emit("import_done", {"importati": importati, "aggiornati": aggiornati, "saltati": saltati})
    log.info(f"Import completato: {importati} nuovi, {aggiornati} aggiornati, {saltati} saltati")
    log.info(f"Colonne trovate: {list(col_map.keys())}")
    return jsonify({"ok": True,
        "msg": f"✅ Import OK: {importati} nuovi, {aggiornati} aggiornati, {saltati} saltati su {tot} totali",
        "importati": importati, "aggiornati": aggiornati, "saltati": saltati})


# ══════════════════════════════════════════════════════════════════════════════
# ASSISTENTE AI — DB cerca i dati, AI formula la risposta
# ══════════════════════════════════════════════════════════════════════════════

import re as _re

def cerca_nel_db(domanda):
    """
    Interpreta la domanda e interroga direttamente il DB.
    Restituisce dati REALI e accurati.
    """
    d = domanda.lower()
    risultati = {}

    # ── DISPONIBILITÀ PEZZO ──────────────────────────────────────────
    # Estrai termini di ricerca (parole > 3 chars che non siano parole comuni)
    stop_words = {'hai','avete','avere','pezzo','pezzi','magazzino','disponibile',
                  'disponibili','quanti','quante','cerco','cerca','serve','hanno',
                  'questo','quella','quello','degli','delle','della','sono','cosa',
                  'come','dove','quando','qual','dici','dirmi','dimmi','voglio'}
    termini = [w for w in _re.findall(r'[a-zA-ZÀ-ÿ]+', d)
               if len(w) > 3 and w not in stop_words]

    if termini:
        # Cerca pezzi che matchano i termini
        conditions = " OR ".join(["(articolo LIKE ? OR marca LIKE ? OR modello LIKE ? OR categoria LIKE ? OR extra1 LIKE ? OR extra2 LIKE ?)"] * len(termini))
        params = []
        for t in termini:
            params.extend([f"%{t}%"] * 6)

        pezzi_trovati = db.fetchall(f"""
            SELECT cmp, articolo, marca, modello, categoria,
                   esistenza, scorta, ubicazione, listino1,
                   cilindrata, carburante, anno_da
            FROM v_giacenza
            WHERE {conditions}
            ORDER BY esistenza DESC
            LIMIT 30
        """, params)
        risultati['pezzi_trovati'] = pezzi_trovati

    # ── PEZZI ESAURITI ───────────────────────────────────────────────
    if any(w in d for w in ['scorta','mancano','esauriti','finiti','ordinare','acquistare']):
        sotto = db.fetchall("""
            SELECT cmp, articolo, marca, ubicazione
            FROM v_giacenza
            WHERE esistenza <= 0
            ORDER BY articolo
            LIMIT 20
        """)
        risultati['sotto_scorta'] = sotto

    # ── PIÙ VENDUTI ───────────────────────────────────────────────────
    periodo = 30
    if any(w in d for w in ['anno','annuale','12 mesi']):
        periodo = 365
    elif any(w in d for w in ['settimana','7 giorni']):
        periodo = 7
    elif any(w in d for w in ['trimestre','3 mesi']):
        periodo = 90

    if any(w in d for w in ['venduti','venduto','scaricati','vendite','vendo','vende',
                             'vendono','richiesti','richiesto','popolare']):
        venduti = db.fetchall(f"""
            SELECT c.descrizione as articolo, c.produttore as marca,
                   SUM(m.quantita) as qty, COUNT(*) as n_vendite
            FROM movimenti_magazzino m
            JOIN componenti c ON c.id = m.componente_id
            WHERE m.tipo = 'scarico'
              AND m.creato_il >= date('now', '-{periodo} days')
            GROUP BY m.componente_id
            ORDER BY qty DESC
            LIMIT 15
        """)
        risultati['piu_venduti'] = venduti
        risultati['periodo_giorni'] = periodo

    # ── PEZZI FERMI ───────────────────────────────────────────────────
    if any(w in d for w in ['fermi','fermo','mesi','venduto','movimentati',
                             'giacciono','stagnanti','invenduti']):
        mesi = 6
        if '12' in d or 'anno' in d: mesi = 12
        elif '3' in d or 'trimestre' in d: mesi = 3

        fermi = db.fetchall(f"""
            SELECT c.codice as cmp, c.descrizione as articolo, c.produttore as marca,
                   v.esistenza,
                   MAX(m.creato_il) as ultimo_movimento
            FROM componenti c
            JOIN v_giacenza v ON v.componente_id = c.id
            LEFT JOIN movimenti_magazzino m ON m.componente_id = c.id
            WHERE c.eliminato = 0 AND v.esistenza > 0
            GROUP BY c.id
            HAVING (ultimo_movimento IS NULL
                    OR ultimo_movimento < date('now', '-{mesi*30} days'))
            ORDER BY v.esistenza DESC
            LIMIT 20
        """)
        risultati['pezzi_fermi'] = fermi
        risultati['mesi_fermi'] = mesi

    # ── STATISTICHE GENERALI ──────────────────────────────────────────
    if any(w in d for w in ['quanti','totale','statistiche','riepilogo',
                             'magazzino','articoli','componenti','valore']):
        stats = db.fetchone("""
            SELECT COUNT(*) as tot,
                   SUM(CASE WHEN esistenza > 0 THEN 1 ELSE 0 END) as disp,
                   SUM(esistenza) as pz_tot
            FROM v_giacenza
        """)
        risultati['statistiche'] = stats

    # ── RECUPERO AUTO ─────────────────────────────────────────────────
    if any(w in d for w in ['recupero','recuperare','smontare','demolire',
                             'arrivata','arrivato','demolizione']):
        # Cerca marca/modello citati nella domanda
        marche_note = ['fiat','ford','volkswagen','vw','golf','opel','renault','peugeot',
                       'citroen','alfa','lancia','bmw','mercedes','audi','toyota','nissan',
                       'hyundai','kia','seat','skoda','volvo','smart','mini','jeep']
        marca_trovata = next((m for m in marche_note if m in d), None)

        if marca_trovata:
            disponibili = db.fetchall("""
                SELECT articolo, esistenza, ubicazione
                FROM v_giacenza
                WHERE LOWER(marca) LIKE ?
                  AND esistenza > 0
                ORDER BY articolo
                LIMIT 40
            """, [f"%{marca_trovata}%"])
            risultati['recupero_auto'] = disponibili
            risultati['marca_recupero'] = marca_trovata

    return risultati


def formatta_risultati(risultati, domanda):
    """Formatta i risultati DB in testo chiaro per l'AI."""
    parti = []

    if risultati.get('pezzi_trovati') is not None:
        pezzi = risultati['pezzi_trovati']
        if pezzi:
            parti.append("PEZZI TROVATI NEL MAGAZZINO (" + str(len(pezzi)) + " risultati):")
            for p in pezzi[:15]:
                es = p['esistenza'] or 0
                stato = "DISPONIBILE" if es > 0 else "ESAURITO"
                riga = f"- {p['articolo']} | {p['marca'] or ''} {p['modello'] or ''} | {stato} | Quantità: {es}"
                if p.get('ubicazione'): riga += f" | Posizione: {p['ubicazione']}"
                if p.get('listino1') and p['listino1'] > 0: riga += f" | Prezzo: €{p['listino1']}"
                if p.get('anno_da'): riga += f" | Anno: {p['anno_da']}"
                parti.append(riga)
        else:
            parti.append("RICERCA NEL MAGAZZINO: Nessun pezzo trovato con questi criteri.")

    if risultati.get('sotto_scorta'):
        sotto = risultati['sotto_scorta']
        parti.append(f"PEZZI SOTTO SCORTA ({len(sotto)} articoli da riordinare):")
        for p in sotto:
            parti.append(f"- {p['articolo']} | {p['marca'] or ''} | Ha: {p['esistenza']} | Minimo: {p['scorta']}")

    if risultati.get('piu_venduti'):
        gg = risultati.get('periodo_giorni', 30)
        periodo_str = f"ultimi {gg} giorni" if gg <= 90 else "ultimo anno"
        venduti = risultati['piu_venduti']
        parti.append(f"PIÙ VENDUTI ({periodo_str}, top {len(venduti)}):")
        for p in venduti:
            parti.append(f"- {p['articolo']} ({p['marca'] or ''}) | {p['qty']} pezzi venduti in {p['n_vendite']} operazioni")

    if risultati.get('pezzi_fermi'):
        mesi = risultati.get('mesi_fermi', 6)
        fermi = risultati['pezzi_fermi']
        parti.append(f"PEZZI FERMI DA {mesi}+ MESI (in magazzino ma non venduti, top {len(fermi)}):")
        for p in fermi:
            ult = (p['ultimo_movimento'] or 'mai')[:10]
            parti.append(f"- {p['articolo']} ({p['marca'] or ''}) | Quantità: {p['esistenza']} | Ultimo movimento: {ult}")

    if risultati.get('valore'):
        v = risultati['valore']
        parti.append(f"VALORE MAGAZZINO (stima su listino):")
        parti.append(f"- Valore totale: €{round(v['valore_totale'] or 0, 2):,.2f}")
        parti.append(f"- Pezzi con prezzo: {v['con_prezzo']} su {v['totale']} disponibili")
        parti.append(f"- Pezzi fisici totali: {v['pezzi_totali']}")

    if risultati.get('statistiche'):
        s = risultati['statistiche']
        parti.append(f"STATISTICHE MAGAZZINO:")
        parti.append(f"- Articoli totali: {s['tot']}")
        parti.append(f"- Disponibili (ES>0): {s['disp']}")
        parti.append(f"- Sotto scorta: {s['sc']}")
        parti.append(f"- Pezzi fisici totali: {s['pz_tot']}")

    if risultati.get('recupero_auto'):
        marca = risultati.get('marca_recupero', '').upper()
        pezzi = risultati['recupero_auto']
        parti.append(f"PEZZI {marca} DISPONIBILI DA RECUPERARE ({len(pezzi)} trovati):")
        for p in pezzi[:20]:
            parti.append(f"- {p['articolo']} | Qtà: {p['esistenza']} | {p['ubicazione'] or ''}")

    return "\n".join(parti) if parti else "Nessun dato trovato per questa ricerca."


@app.route("/api/assistente", methods=["POST"])
@require_login
def api_assistente():
    import urllib.request, json as _j

    d       = request.json or {}
    domanda = (d.get("domanda") or "").strip()
    if not domanda:
        return jsonify({"ok": False, "msg": "Domanda vuota"}), 400

    # 1. Cerca dati REALI nel database
    try:
        risultati = cerca_nel_db(domanda)
        dati_reali = formatta_risultati(risultati, domanda)
    except Exception as e:
        log.error(f"Errore ricerca DB: {e}")
        dati_reali = "Errore nel recupero dati."

    # 2. Usa AI solo per formulare risposta naturale sui dati reali
    system_prompt = """Sei PERI, assistente del magazzino PerilCar.
Ti vengono forniti dati REALI estratti dal database. Usali per rispondere.
REGOLE ASSOLUTE:
- Usa SOLO i dati forniti, non inventare nulla
- Se i dati dicono "ESAURITO" di lo chiaramente
- Se i dati dicono "DISPONIBILE" con quantità X, riporta esattamente X
- Rispondi in italiano, max 5 frasi, sii diretto e preciso
- Non aggiungere informazioni che non sono nei dati forniti"""

    user_msg = f"""Domanda: {domanda}

DATI REALI DAL DATABASE:
{dati_reali}

Rispondi basandoti ESCLUSIVAMENTE sui dati sopra."""

    # Prova a usare Ollama con timeout breve (30s)
    # Se non risponde, restituisce comunque i dati DB
    try:
        payload = _j.dumps({
            "model": "llama3.2:latest",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_msg}
            ],
            "stream": False,
            "options": {"temperature": 0.1, "num_predict": 200, "num_ctx": 2048}
        }).encode()

        req = urllib.request.Request(
            "http://localhost:11434/api/chat",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = _j.loads(resp.read())
            risposta = result["message"]["content"]
            return jsonify({"ok": True, "risposta": risposta})
    except Exception:
        # Fallback: restituisce dati DB formattati in HTML leggibile
        pass

    # Ollama non disponibile: formatta i dati DB in modo leggibile
    if not dati_reali or dati_reali == "Nessun dato trovato per questa ricerca.":
        risposta_html = "❌ Nessun risultato trovato nel magazzino per questa ricerca."
    else:
        # Converti il plain text in HTML con evidenziazione
        lines = dati_reali.split("\n")
        parts = []
        for line in lines:
            if not line.strip():
                continue
            if line.isupper() or line.endswith(":"):
                parts.append(f"<b style='color:#e94c00'>{line}</b>")
            elif line.startswith("- "):
                parts.append(f"<span style='display:block;padding:1px 0 1px 8px;border-left:2px solid #dde1e8'>{line[2:]}</span>")
            else:
                parts.append(line)
        risposta_html = "<br>".join(parts)
        risposta_html = f"<small style='color:#9ca3af;font-style:italic'>⚙️ Risposta diretta dal DB (Ollama non attivo)</small><br><br>{risposta_html}"

    return jsonify({"ok": True, "risposta": risposta_html})


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# DEMOLIZIONI — usa dem (db/demolizioni.db) separato da magazzino
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/demolizioni")
@require_login
def page_demolizioni():
    return render_template("demolizioni.html", user=cu())

@app.route("/demolisci")
@require_login
def page_demolisci():
    return render_template("demolisci.html", user=cu())

# ── ANAGRAFICHE ───────────────────────────────────────────────────────────────

@app.route("/api/anagrafiche", methods=["GET"])
@require_login
def api_anagrafiche():
    rows = dem.all("SELECT id, nominativo, cf_piva, telefono, email FROM anagrafiche ORDER BY nominativo")
    return jsonify(rows)

@app.route("/api/anagrafiche", methods=["POST"])
@require_login
def api_anagrafiche_crea():
    d = request.json or {}
    try:
        campi = ["nominativo","cognome","nome","cf_piva","sesso","tipo_societa","tipo",
                 "data_nascita","luogo_nascita","prov_nascita","comune","provincia",
                 "via","civico","cap","tipo_doc","num_doc","data_doc","rilasciato_da",
                 "telefono","cellulare","fax","email","indirizzo","note"]
        cols = [f for f in campi if d.get(f) is not None]
        vals = [d[f] for f in cols]
        nid  = dem.run("INSERT INTO anagrafiche ("+",".join(cols)+") VALUES ("+",".join(["?"]*len(cols))+")", vals)
        return jsonify({"ok": True, "id": nid, "msg": "Salvata", "nominativo": d.get("nominativo","")})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/anagrafiche/<int:aid>", methods=["GET"])
@require_login
def api_anagrafica_get(aid):
    row = dem.one("SELECT * FROM anagrafiche WHERE id=?", (aid,))
    if not row: return jsonify({"ok": False}), 404
    return jsonify({"ok": True, "data": row})

@app.route("/api/anagrafiche/<int:aid>", methods=["PUT"])
@require_login
def api_anagrafica_update(aid):
    d = request.json or {}
    try:
        campi = ["nominativo","cognome","nome","cf_piva","sesso","tipo_societa","tipo",
                 "data_nascita","luogo_nascita","prov_nascita","comune","provincia",
                 "via","civico","cap","tipo_doc","num_doc","data_doc","rilasciato_da",
                 "telefono","cellulare","fax","email","indirizzo","note"]
        cols = [f for f in campi if f in d]
        vals = [d[f] for f in cols] + [aid]
        if cols:
            dem.run("UPDATE anagrafiche SET "+",".join(f+"=?" for f in cols)+" WHERE id=?", vals)
        return jsonify({"ok": True, "id": aid, "msg": "Aggiornata"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/anagrafiche/<int:aid>/allegati", methods=["GET"])
@require_login
def api_anagrafica_allegati(aid):
    import os
    d = os.path.join(os.path.dirname(__file__), "uploads", "anagrafiche", str(aid))
    if not os.path.exists(d): return jsonify([])
    return jsonify([{"nome": f, "size": os.path.getsize(os.path.join(d,f)),
                     "url": f"/uploads/anagrafiche/{aid}/{f}"} for f in os.listdir(d)])

@app.route("/api/anagrafiche/<int:aid>/allegati", methods=["POST"])
@require_login
def api_anagrafica_upload(aid):
    import os
    d = os.path.join(os.path.dirname(__file__), "uploads", "anagrafiche", str(aid))
    os.makedirs(d, exist_ok=True)
    saved = []
    for f in request.files.getlist("files"):
        if f.filename:
            name = f.filename.replace("/","_").replace("\\","_")
            f.save(os.path.join(d, name)); saved.append(name)
    return jsonify({"ok": True, "files": saved})

# ── VEICOLI ───────────────────────────────────────────────────────────────────

@app.route("/api/veicoli", methods=["GET"])
@require_login
def api_veicoli():
    return jsonify(dem.all("SELECT * FROM veicoli ORDER BY id DESC"))

@app.route("/api/veicoli/<int:vid>", methods=["GET"])
@require_login
def api_veicolo_get(vid):
    row = dem.one("SELECT * FROM veicoli WHERE id=?", (vid,))
    if not row: return jsonify({"ok": False}), 404
    return jsonify({"ok": True, "data": row})

@app.route("/api/veicoli", methods=["POST"])
@require_login
def api_veicoli_crea():
    d = request.json or {}
    try:
        campi = ["targa","telaio","classe","marca","modello","anno_immatricolazione","num_motore","colore","note"]
        cols  = [f for f in campi if d.get(f) is not None]
        vals  = [d[f] for f in cols]
        nid   = dem.run("INSERT INTO veicoli ("+",".join(cols)+") VALUES ("+",".join(["?"]*len(cols))+")", vals)
        marca = d.get("marca",""); modello = d.get("modello",""); targa = d.get("targa","")
        label = (marca+" "+modello).strip() + (" ("+targa+")" if targa else "")
        return jsonify({"ok": True, "id": nid, "msg": "Veicolo salvato", "label": label})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/veicoli/<int:vid>", methods=["PUT"])
@require_login
def api_veicolo_update(vid):
    d = request.json or {}
    try:
        campi = ["targa","telaio","classe","marca","modello","anno_immatricolazione","num_motore","colore","note"]
        cols  = [f for f in campi if f in d]
        if cols:
            dem.run("UPDATE veicoli SET "+",".join(f+"=?" for f in cols)+" WHERE id=?",
                    [d[f] for f in cols]+[vid])
        return jsonify({"ok": True, "id": vid})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/veicoli/<int:vid>/allegati", methods=["GET"])
@require_login
def api_veicolo_allegati(vid):
    import os
    d = os.path.join(os.path.dirname(__file__), "uploads", "veicoli", str(vid))
    if not os.path.exists(d): return jsonify([])
    return jsonify([{"nome": f, "size": os.path.getsize(os.path.join(d,f)),
                     "url": f"/uploads/veicoli/{vid}/{f}"} for f in os.listdir(d)])

@app.route("/api/veicoli/<int:vid>/allegati", methods=["POST"])
@require_login
def api_veicolo_upload(vid):
    import os
    d = os.path.join(os.path.dirname(__file__), "uploads", "veicoli", str(vid))
    os.makedirs(d, exist_ok=True)
    saved = []
    for f in request.files.getlist("files"):
        if f.filename:
            name = f.filename.replace("/","_").replace("\\","_")
            f.save(os.path.join(d, name)); saved.append(name)
    return jsonify({"ok": True, "files": saved})

@app.route("/uploads/<path:filename>")
@require_login
def serve_upload(filename):
    import os
    from flask import send_from_directory
    return send_from_directory(os.path.join(os.path.dirname(__file__), "uploads"), filename)

# ── DEMOLIZIONI ───────────────────────────────────────────────────────────────

def _dem_query(q="", params=()):
    sql = """SELECT d.id, d.data_presa_in_carico, d.ora_presa_in_carico,
               d.reg_demolitori, d.pag_reg, d.veicolo_id, d.proprietario_id, d.detentore_id,
               d.ufficio_provinciale, d.targhe_consegnate, d.carta_circolazione,
               d.concessionaria, d.peso_effettivo_kg, d.peso_netto_kg,
               d.modalita_radiazione, d.num_albatros, d.note, COALESCE(d.primo_trattamento,0) AS primo_trattamento,
               COALESCE(v.marca||' '||v.modello||' ('||v.targa||')','') AS veicolo_str,
               COALESCE(p.nominativo,'') AS proprietario_str,
               COALESCE(det.nominativo,'') AS detentore_str
             FROM demolizioni d
             LEFT JOIN veicoli v ON v.id=d.veicolo_id
             LEFT JOIN anagrafiche p ON p.id=d.proprietario_id
             LEFT JOIN anagrafiche det ON det.id=d.detentore_id"""
    if q:
        sql += " WHERE " + q
    sql += " ORDER BY d.id DESC"
    return dem.all(sql, params)

@app.route("/api/demolizioni/cerca")
@require_login
def api_demolizioni_cerca():
    return jsonify(_dem_query())

@app.route("/api/demolizioni/prossimi-progressivi")
@require_login
def api_demolizioni_progressivi():
    import datetime
    nid  = (dem.one("SELECT COALESCE(MAX(id),0)+1 AS v FROM demolizioni") or {}).get("v",1)
    npag = (dem.one("SELECT COALESCE(MAX(CAST(pag_reg AS INTEGER)),0)+1 AS v FROM demolizioni") or {}).get("v",1)
    now  = datetime.datetime.now()
    return jsonify({"ok": True, "next_id": nid, "next_pag": npag,
                    "reg_demolitori": f"01/{now.year}",
                    "data": now.strftime("%Y-%m-%d"), "ora": now.strftime("%H:%M")})

@app.route("/api/demolizioni", methods=["POST"])
@require_login
def api_demolizioni_crea():
    u  = cu(); d = request.json or {}
    campi = ["data_presa_in_carico","ora_presa_in_carico","reg_demolitori","pag_reg",
             "veicolo_id","proprietario_id","detentore_id","ufficio_provinciale",
             "targhe_consegnate","carta_circolazione","concessionaria",
             "peso_effettivo_kg","peso_netto_kg","modalita_radiazione",
             "num_albatros","certificato_id","note","creato_da"]
    vals  = [d.get(f) for f in campi[:-1]] + [u.get("id")]
    try:
        nid = dem.run("INSERT INTO demolizioni ("+",".join(campi)+") VALUES ("+",".join(["?"]*len(campi))+")", vals)
        return jsonify({"ok": True, "id": nid, "msg": "Demolizione salvata"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/demolizioni/<int:did>", methods=["PUT"])
@require_login
def api_demolizioni_update(did):
    d = request.json or {}
    campi = ["data_presa_in_carico","ora_presa_in_carico","reg_demolitori","pag_reg",
             "veicolo_id","proprietario_id","detentore_id","ufficio_provinciale",
             "targhe_consegnate","carta_circolazione","concessionaria",
             "peso_effettivo_kg","peso_netto_kg","modalita_radiazione","note","primo_trattamento"]
    cols = [f for f in campi if f in d]
    if cols:
        dem.run("UPDATE demolizioni SET "+",".join(f+"=?" for f in cols)+" WHERE id=?",
                [d[f] for f in cols]+[did])
    return jsonify({"ok": True, "msg": "Aggiornata"})

@app.route("/api/demolizioni/<int:did>", methods=["DELETE"])
@require_login
def api_demolizioni_delete(did):
    dem.run("DELETE FROM ricambi_sottratti WHERE demolizione_id=?", (did,))
    dem.run("DELETE FROM demolizioni WHERE id=?", (did,))
    return jsonify({"ok": True})

@app.route("/api/demolizioni/<int:did>/ricambi", methods=["GET"])
@require_login
def api_ricambi_get(did):
    return jsonify(dem.all("SELECT * FROM ricambi_sottratti WHERE demolizione_id=? ORDER BY id", (did,)))

@app.route("/api/demolizioni/<int:did>/ricambi", methods=["POST"])
@require_login
def api_ricambi_add(did):
    d    = request.json or {}
    nome = d.get("pezzo_nome") or ""
    peso = d.get("peso_kg", 0)
    dem.run("INSERT INTO ricambi_sottratti (demolizione_id,componente_id,peso_kg,note,pezzo_nome) VALUES (?,?,?,?,?)",
            (did, None, peso, d.get("note",""), nome))
    # Ricalcola peso netto
    tot = (dem.one("SELECT COALESCE(SUM(peso_kg),0) AS t FROM ricambi_sottratti WHERE demolizione_id=?", (did,)) or {}).get("t",0)
    dem.run("UPDATE demolizioni SET peso_netto_kg = MAX(0, COALESCE(peso_effettivo_kg,0)-?) WHERE id=?", (tot, did))
    return jsonify({"ok": True})

@app.route("/api/demolizioni/ricambi/<int:rid>", methods=["DELETE"])
@require_login
def api_ricambi_del(rid):
    row = dem.one("SELECT demolizione_id FROM ricambi_sottratti WHERE id=?", (rid,))
    dem.run("DELETE FROM ricambi_sottratti WHERE id=?", (rid,))
    if row:
        did = row["demolizione_id"]
        tot = (dem.one("SELECT COALESCE(SUM(peso_kg),0) AS t FROM ricambi_sottratti WHERE demolizione_id=?", (did,)) or {}).get("t",0)
        dem.run("UPDATE demolizioni SET peso_netto_kg = MAX(0, COALESCE(peso_effettivo_kg,0)-?) WHERE id=?", (tot, did))
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# SCHEDE DEMOLIZIONE — salvataggio scheda /demolisci nel DB
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/schede-demolizione", methods=["POST"])
@require_login
def api_scheda_salva():
    u = cu(); d = request.json or {}
    dem_ids   = d.get("dem_ids", "")      # es. "3,7,12"
    data_tratt = d.get("data_trattamento")
    righe_json = d.get("righe_json", "[]")  # JSON serializzato
    peso_eff   = d.get("peso_eff_tot", 0)
    peso_ric   = d.get("peso_ricambi_tot", 0)
    peso_netto = d.get("peso_netto_tot", 0)
    if not dem_ids:
        return jsonify({"ok": False, "msg": "dem_ids obbligatorio"}), 400
    try:
        # Upsert: se esiste già una scheda per questi dem_ids, aggiorna
        existing = dem.one("SELECT id FROM schede_demolizione WHERE dem_ids=?", (dem_ids,))
        if existing:
            dem.run("""UPDATE schede_demolizione
                SET data_trattamento=?, peso_eff_tot=?, peso_ricambi_tot=?, peso_netto_tot=?,
                    righe_json=?, modificato_il=datetime('now')
                WHERE id=?""",
                (data_tratt, peso_eff, peso_ric, peso_netto, righe_json, existing["id"]))
            return jsonify({"ok": True, "msg": "Scheda aggiornata nel DB", "id": existing["id"], "action": "update"})
        else:
            nid = dem.run("""INSERT INTO schede_demolizione
                (dem_ids, data_trattamento, peso_eff_tot, peso_ricambi_tot, peso_netto_tot, righe_json, creato_da)
                VALUES (?,?,?,?,?,?,?)""",
                (dem_ids, data_tratt, peso_eff, peso_ric, peso_netto, righe_json, u.get("id")))
            return jsonify({"ok": True, "msg": "Scheda salvata nel DB", "id": nid, "action": "create"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/schede-demolizione/<dem_ids>", methods=["GET"])
@require_login
def api_scheda_get(dem_ids):
    row = dem.one("SELECT * FROM schede_demolizione WHERE dem_ids=? ORDER BY id DESC LIMIT 1", (dem_ids,))
    if not row:
        return jsonify({"ok": False, "found": False})
    return jsonify({"ok": True, "found": True, "data": row})

@app.route("/api/demolizioni/<int:did>", methods=["GET"])
@require_login
def api_demolizione_get(did):
    row = _dem_query(f"d.id={did}")
    if not row:
        return jsonify({"ok": False, "msg": "Non trovata"}), 404
    return jsonify({"ok": True, "data": row[0]})

# ══════════════════════════════════════════════════════════════════════════════
# GESTIONE UTENTI (solo admin)
# ══════════════════════════════════════════════════════════════════════════════

def require_admin(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = cu()
        if not u:
            return jsonify({"ok": False, "msg": "Non autenticato"}), 401
        if u.get("ruolo") != "admin":
            return jsonify({"ok": False, "msg": "Accesso negato — solo admin"}), 403
        return f(*args, **kwargs)
    return wrapper

@app.route("/utenti")
@require_login
def page_utenti():
    if cu().get("ruolo") != "admin":
        return redirect("/dashboard")
    return render_template("utenti.html", user=cu())

@app.route("/api/utenti", methods=["GET"])
@require_admin
def api_utenti_lista():
    try:
        ucols = [r["name"] for r in db.fetchall("PRAGMA table_info(utenti)")]
        sel   = "id,username,ruolo,attivo,creato_il"
        if "nome_completo" in ucols: sel += ",nome_completo"
        where = "WHERE eliminato=0" if "eliminato" in ucols else ""
        rows  = db.fetchall(f"SELECT {sel} FROM utenti {where} ORDER BY username")
        # Normalizza: aggiungi nome_completo se manca
        for r in rows:
            if "nome_completo" not in r.keys():
                r = dict(r); r["nome_completo"] = ""
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        log.warning(f"api_utenti_lista error: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/utenti", methods=["POST"])
@require_admin
def api_utenti_crea():
    import hashlib
    d = request.json or {}
    username = (d.get("username") or "").strip().lower()
    password = (d.get("password") or "").strip()
    ruolo    = d.get("ruolo", "operatore")
    nome     = d.get("nome_completo", "").strip()
    if not username or not password:
        return jsonify({"ok": False, "msg": "Username e password obbligatori"}), 400
    if len(password) < 6:
        return jsonify({"ok": False, "msg": "Password minimo 6 caratteri"}), 400
    if db.fetchone("SELECT id FROM utenti WHERE username=? AND eliminato=0", (username,)):
        return jsonify({"ok": False, "msg": f"Username '{username}' già esistente"}), 400
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    try:
        with db._write_lock:
            conn = db.get_connection()
            try:
                # Adatta INSERT alle colonne disponibili
                ucols2 = [r[1] for r in conn.execute("PRAGMA table_info(utenti)").fetchall()]
                fields = ["username","password_hash","ruolo","attivo"]
                values = [username, pwd_hash, ruolo, 1]
                if "nome_completo" in ucols2 and nome:
                    fields.append("nome_completo"); values.append(nome)
                if "eliminato" in ucols2:
                    fields.append("eliminato"); values.append(0)
                sql = "INSERT INTO utenti("+",".join(fields)+") VALUES("+",".join(["?"]*len(fields))+")"
                cur = conn.execute(sql, values)
                uid = cur.lastrowid
                conn.execute("INSERT INTO log_operazioni(utente_id,username,modulo,azione,tabella,record_id) VALUES(?,?,?,?,?,?)",
                    (cu().get("id"), cu().get("username"), "UTENTI", "CREA", "utenti", uid))
                conn.commit()
            finally:
                conn.close()
        return jsonify({"ok": True, "msg": f"Utente '{username}' creato", "id": uid})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/utenti/<int:uid>", methods=["PUT"])
@require_admin
def api_utenti_modifica(uid):
    import hashlib
    d = request.json or {}
    updates = []
    vals = []
    if "ruolo" in d:
        updates.append("ruolo=?"); vals.append(d["ruolo"])
    if "nome_completo" in d:
        # Aggiungi colonna se non esiste
        try:
            ucols_m = [r[1] for r in db.fetchall("PRAGMA table_info(utenti)")]
            if "nome_completo" not in ucols_m:
                db_write([("ALTER TABLE utenti ADD COLUMN nome_completo TEXT", ())])
        except: pass
        updates.append("nome_completo=?"); vals.append(d["nome_completo"])
    if "attivo" in d:
        updates.append("attivo=?"); vals.append(1 if d["attivo"] else 0)
    if d.get("password"):
        if len(d["password"]) < 6:
            return jsonify({"ok": False, "msg": "Password minimo 6 caratteri"}), 400
        updates.append("password_hash=?")
        vals.append(hashlib.sha256(d["password"].encode()).hexdigest())
    if not updates:
        return jsonify({"ok": False, "msg": "Nessun campo da aggiornare"}), 400
    vals.append(uid)
    try:
        db_write([(f"UPDATE utenti SET {','.join(updates)} WHERE id=? AND eliminato=0", vals),
                  ("INSERT INTO log_operazioni(utente_id,username,modulo,azione,tabella,record_id) VALUES(?,?,?,?,?,?)",
                   (cu().get("id"), cu().get("username"), "UTENTI", "MODIFICA", "utenti", uid))])
        return jsonify({"ok": True, "msg": "Utente aggiornato"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500
@app.route("/api/utenti/<int:uid>", methods=["DELETE"])
@require_admin
def api_utenti_elimina(uid):
    if uid == cu().get("id"):
        return jsonify({"ok": False, "msg": "Non puoi eliminare te stesso"}), 400
    try:
        ucols_e = [r[1] for r in db.fetchall("PRAGMA table_info(utenti)")]
        if "eliminato" in ucols_e:
            db_write([("UPDATE utenti SET eliminato=1 WHERE id=?", (uid,)),
                      ("INSERT INTO log_operazioni(utente_id,username,modulo,azione,tabella,record_id) VALUES(?,?,?,?,?,?)",
                       (cu().get("id"), cu().get("username"), "UTENTI", "ELIMINA", "utenti", uid))])
        else:
            db_write([("DELETE FROM utenti WHERE id=?", (uid,))])
        return jsonify({"ok": True, "msg": "Utente eliminato"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/utenti/cambia-password", methods=["POST"])
@require_login
def api_cambia_password():
    """Ogni utente può cambiare la propria password."""
    import hashlib
    d = request.json or {}
    vecchia   = d.get("vecchia", "")
    nuova     = d.get("nuova", "")
    if len(nuova) < 6:
        return jsonify({"ok": False, "msg": "Nuova password minimo 6 caratteri"}), 400
    u = cu()
    utente = db.fetchone("SELECT password_hash FROM utenti WHERE id=? AND eliminato=0", (u.get("id"),))
    if not utente:
        return jsonify({"ok": False, "msg": "Utente non trovato"}), 404
    if utente["password_hash"] != hashlib.sha256(vecchia.encode()).hexdigest():
        return jsonify({"ok": False, "msg": "Password attuale errata"}), 400
    db_write([("UPDATE utenti SET password_hash=? WHERE id=?",
               (hashlib.sha256(nuova.encode()).hexdigest(), u.get("id")))])
    return jsonify({"ok": True, "msg": "Password aggiornata"})


# ══════════════════════════════════════════════════════════════════════════════
# MODULO INVENTARIO — PC controller + Mobile client
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/inventario")
@require_login
def page_inventario():
    return render_template("inventario.html", user=cu())

@app.route("/inventario-mobile")
@require_login
def page_inventario_mobile():
    sid = request.args.get("sessione")
    return render_template("inventario_mobile.html", user=cu(), sessione_id=sid or "")

# ── CATEGORIE disponibili ─────────────────────────────────────────────
@app.route("/api/inventario/categorie")
@require_login
def api_inv_categorie():
    rows = db.fetchall("""
        SELECT categoria, COUNT(*) as n, SUM(esistenza) as tot_pezzi
        FROM v_giacenza
        WHERE categoria IS NOT NULL AND categoria != ''
        GROUP BY categoria ORDER BY categoria
    """)
    return jsonify(rows)

@app.route("/api/inventario/descrizioni")
@require_login
def api_inv_descrizioni():
    """Cerca descrizioni/articoli nel magazzino per autocomplete."""
    q = request.args.get("q","").strip()
    if len(q) < 1:
        return jsonify([])
    rows = db.fetchall("""
        SELECT articolo as descrizione, COUNT(*) as n, SUM(esistenza) as tot_pezzi
        FROM v_giacenza
        WHERE articolo LIKE ? AND articolo IS NOT NULL AND articolo != ''
        GROUP BY articolo ORDER BY n DESC, articolo LIMIT 40
    """, (f"%{q}%",))
    return jsonify(rows)

# ── SESSIONI ──────────────────────────────────────────────────────────
@app.route("/api/inventario/sessioni")
@require_login
def api_inv_sessioni():
    rows = db.fetchall("""
        SELECT s.*, u.username as creato_da_nome
        FROM sessioni_inventario s
        LEFT JOIN utenti u ON u.id = s.creato_da
        ORDER BY s.id DESC LIMIT 50
    """)
    return jsonify(rows)

@app.route("/api/inventario/sessioni", methods=["POST"])
@require_login
def api_inv_crea_sessione():
    import json
    d  = request.json or {}
    u  = cu()
    categoria  = d.get("categoria", "").strip()
    descrizione= d.get("descrizione", "").strip()
    nome       = d.get("nome", descrizione or categoria or "Inventario").strip()
    filtro     = descrizione or categoria
    if not filtro:
        return jsonify({"ok": False, "msg": "Seleziona una descrizione"}), 400

    try:
        # Recupera componenti PRIMA di acquisire il write lock (evita deadlock)
        if descrizione:
            comps = db.fetchall(
                "SELECT componente_id, esistenza FROM v_giacenza WHERE articolo=? ORDER BY articolo",
                (descrizione,))
        else:
            comps = db.fetchall(
                "SELECT componente_id, esistenza FROM v_giacenza WHERE categoria=? ORDER BY articolo",
                (categoria,))

        if not comps:
            return jsonify({"ok": False, "msg": f"Nessun componente trovato per '{filtro}'"}), 400

        with db._write_lock:
            conn = db.get_connection()
            conn.execute("PRAGMA synchronous=OFF")
            try:
                # Crea sessione
                cur = conn.execute(
                    "INSERT INTO sessioni_inventario(nome,categoria,stato,creato_da) VALUES(?,?,?,?)",
                    (nome, filtro, "attiva", u.get("id")))
                sid = cur.lastrowid

                # Insert bulk
                conn.executemany(
                    "INSERT INTO inventario_righe(sessione_id,componente_id,qty_attesa,stato,ordine) VALUES(?,?,?,?,?)",
                    [(sid, c["componente_id"], int(c["esistenza"] or 0), "sospeso", i)
                     for i, c in enumerate(comps)])

                conn.execute(
                    "UPDATE sessioni_inventario SET totale_pezzi=? WHERE id=?",
                    (len(comps), sid))
                conn.commit()
                log.info(f"Inventario sessione {sid} creata: {len(comps)} pezzi categoria '{categoria}'")

                # Notifica tutti i client connessi
                socketio.emit("inventario_nuova_sessione", {
                    "sessione_id": sid, "nome": nome, "categoria": categoria,
                    "totale": len(comps)
                })
                return jsonify({"ok": True, "id": sid, "totale": len(comps), "msg": f"Sessione creata: {len(comps)} pezzi"})
            except Exception as e:
                conn.rollback()
                raise
            finally:
                conn.close()
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/inventario/sessioni/<int:sid>")
@require_login
def api_inv_sessione_dettaglio(sid):
    sess = db.fetchone("SELECT * FROM sessioni_inventario WHERE id=?", (sid,))
    if not sess:
        return jsonify({"ok": False, "msg": "Sessione non trovata"}), 404
    righe = db.fetchall("""
        SELECT r.*,
               c.codice          AS cmp,
               c.descrizione     AS articolo,
               c.produttore      AS marca,
               c.modello,
               c.anno            AS anno_da,
               c.alimentazione   AS carburante,
               c.colore,
               c.cilindrata,
               c.ubicazione,
               c.cod_barre,
               c.udm             AS cod_udm,
               c.note,
               c.extra3,
               c.cod_prod_forn,
               c.cod_fornitore,
               c.fornitore,
               c.immagine        AS immagine_path
        FROM inventario_righe r
        JOIN componenti c ON c.id = r.componente_id
        WHERE r.sessione_id=?
        ORDER BY r.ordine
    """, (sid,))
    return jsonify({"ok": True, "sessione": dict(sess), "righe": righe})

@app.route("/api/inventario/sessioni/<int:sid>/prossimo")
@require_login
def api_inv_prossimo(sid):
    """Restituisce il prossimo pezzo sospeso da inventariare."""
    riga = db.fetchone("""
        SELECT r.*, c.codice as cmp, c.descrizione as articolo,
               c.produttore as marca, c.immagine as immagine_path,
               c.colore, c.modello, c.anno as anno_da,
               c.cilindrata, c.alimentazione as carburante,
               c.udm as cod_udm,
               c.cod_barre, c.ubicazione,
               c.fornitore, c.cod_fornitore, c.note,
               c.extra3, c.cod_prod_forn
        FROM inventario_righe r
        JOIN componenti c ON c.id = r.componente_id
        WHERE r.sessione_id=? AND r.stato='sospeso'
        ORDER BY r.ordine LIMIT 1
    """, (sid,))
    if not riga:
        # Controlla se ci sono rimandati
        rimandato = db.fetchone("""
            SELECT r.*, c.codice as cmp, c.descrizione as articolo,
                   c.produttore as marca, c.immagine as immagine_path,
                   c.colore, c.modello, c.anno as anno_da,
                   c.cilindrata, c.alimentazione as carburante,
                   c.udm as cod_udm,
                   c.cod_barre, c.ubicazione,
                   c.fornitore, c.cod_fornitore, c.note,
                   c.extra3, c.cod_prod_forn
            FROM inventario_righe r
            JOIN componenti c ON c.id = r.componente_id
            WHERE r.sessione_id=? AND r.stato='rimandato'
            ORDER BY r.ordine LIMIT 1
        """, (sid,))
        if rimandato:
            return jsonify({"ok": True, "riga": dict(rimandato), "fase": "rimandati"})
        return jsonify({"ok": True, "riga": None, "fase": "completato"})
    return jsonify({"ok": True, "riga": dict(riga), "fase": "principale"})

@app.route("/api/inventario/sessioni/<int:sid>/stats")
@require_login
def api_inv_stats(sid):
    stats = db.fetchone("""
        SELECT
            COUNT(*) as totale,
            SUM(CASE WHEN stato='confermato' THEN 1 ELSE 0 END) as confermati,
            SUM(CASE WHEN stato='mancante' THEN 1 ELSE 0 END) as mancanti,
            SUM(CASE WHEN stato='rimandato' THEN 1 ELSE 0 END) as rimandati,
            SUM(CASE WHEN stato='sospeso' THEN 1 ELSE 0 END) as sospesi
        FROM inventario_righe WHERE sessione_id=?
    """, (sid,))
    return jsonify(dict(stats) if stats else {})

@app.route("/api/inventario/righe/<int:rid>", methods=["PUT"])
@require_login
def api_inv_aggiorna_riga(rid):
    """Aggiorna stato di una riga inventario (dal mobile o dal PC)."""
    import json
    d     = request.json or {}
    u     = cu()
    stato = d.get("stato")           # confermato | mancante | rimandato
    qty   = d.get("qty_trovata")
    note  = d.get("note", "")

    # Leggi riga PRIMA del write lock
    riga = db.fetchone("SELECT * FROM inventario_righe WHERE id=?", (rid,))
    if not riga:
        return jsonify({"ok": False, "msg": "Riga non trovata"}), 404

    # Pre-calcola valori
    sid  = riga["sessione_id"]
    cid  = riga["componente_id"]
    qatt = riga["qty_attesa"] or 0
    qtrv = qty if qty is not None else qatt

    try:
        with db._write_lock:
            conn = db.get_connection()
            try:
                # Aggiorna riga inventario
                conn.execute("""UPDATE inventario_righe
                    SET stato=?, qty_trovata=?, note=?, aggiornato_il=datetime('now')
                    WHERE id=?""",
                    (stato, qtrv, note, rid))

                # Se confermato o mancante: registra movimento se qty diversa
                if stato in ("confermato", "mancante"):
                    diff = qtrv - qatt
                    if diff != 0:
                        tipo_mov = "carico" if diff > 0 else "scarico"
                        q_mov    = abs(diff)
                        mov_cols = [r[1] for r in conn.execute("PRAGMA table_info(movimenti_magazzino)").fetchall()]
                        m_fields = ["componente_id","tipo","quantita","riferimento","note","utente_id"]
                        m_vals   = [cid, tipo_mov, q_mov, f"INVENTARIO-{sid}", f"Rettifica inventario: trovati {qtrv} attesi {qatt}", u.get("id")]
                        if "quantita_prima" in mov_cols:
                            m_fields += ["quantita_prima","quantita_dopo"]
                            m_vals   += [qatt, qtrv]
                        conn.execute(
                            "INSERT INTO movimenti_magazzino ("+",".join(m_fields)+") VALUES ("+",".join(["?"]*len(m_fields))+")",
                            m_vals)
                        # Aggiorna giacenza denormalizzata + invalida cache
                        conn.execute("""
                            UPDATE componenti SET giacenza = COALESCE((
                                SELECT SUM(CASE
                                    WHEN tipo IN ('carico','inventario') THEN  quantita
                                    WHEN tipo = 'scarico'               THEN -quantita
                                    WHEN tipo = 'rettifica'             THEN  quantita
                                    ELSE 0 END)
                                FROM movimenti_magazzino WHERE componente_id = ?
                            ), 0) WHERE id = ?""", (cid, cid))
                        _comp_cache["ts"] = 0

                # Aggiorna contatori sessione
                stats = conn.execute("""
                    SELECT
                        SUM(CASE WHEN stato='confermato' THEN 1 ELSE 0 END) as c,
                        SUM(CASE WHEN stato='mancante'   THEN 1 ELSE 0 END) as m,
                        SUM(CASE WHEN stato='rimandato'  THEN 1 ELSE 0 END) as r
                    FROM inventario_righe WHERE sessione_id=?
                """, (sid,)).fetchone()
                conn.execute("""UPDATE sessioni_inventario
                    SET confermati=?, mancanti=?, sospesi=?
                    WHERE id=?""",
                    (stats[0] or 0, stats[1] or 0, stats[2] or 0, sid))

                # Log
                try:
                    conn.execute("""INSERT INTO log_operazioni
                        (utente_id,username,modulo,azione,tabella,record_id,dati_nuovi)
                        VALUES(?,?,?,?,?,?,?)""",
                        (u.get("id"),u.get("username"),"INVENTARIO",stato.upper(),
                         "inventario_righe",rid,str({"qty":qtrv,"stato":stato})))
                except: pass

                conn.commit()
            except Exception as e:
                conn.rollback(); raise
            finally:
                conn.close()

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

    # Notifica FUORI dal lock — evita blocking su threading mode
    socketio.emit("inventario_aggiornamento", {
        "sessione_id": sid, "riga_id": rid,
        "componente_id": cid, "stato": stato,
        "qty_trovata": qtrv, "qty_attesa": qatt
    }, room=f"inventario_{sid}")
    return jsonify({"ok": True, "msg": f"Pezzo {stato}"})

@app.route("/api/inventario/righe/<int:rid>/foto", methods=["POST"])
@require_login
def api_inv_foto(rid):
    """Upload foto da mobile per un pezzo inventario."""
    import uuid
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f   = request.files["file"]
    ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else "jpg"
    if ext not in {"png","jpg","jpeg","webp","gif"}:
        return jsonify({"ok": False, "msg": "Solo immagini"}), 400

    riga = db.fetchone("SELECT * FROM inventario_righe WHERE id=?", (rid,))
    if not riga:
        return jsonify({"ok": False, "msg": "Riga non trovata"}), 404

    fname = f"inv_{rid}_{uuid.uuid4().hex[:8]}.{ext}"
    f.save(UPLOAD_FOLDER / fname)
    url   = f"/static/uploads/{fname}"

    cid = riga["componente_id"]
    try:
        with db._write_lock:
            conn = db.get_connection()
            try:
                # Salva url foto nella riga inventario (append, non sovrascrive)
                old_riga = conn.execute("SELECT foto_url FROM inventario_righe WHERE id=?", (rid,)).fetchone()
                old_foto_url = old_riga[0] if old_riga and old_riga[0] else ""
                new_foto_url = (old_foto_url + "|" + url).strip("|") if old_foto_url else url
                conn.execute("UPDATE inventario_righe SET foto_url=? WHERE id=?", (new_foto_url, rid))
                # Aggiorna immagine del componente se non ne ha già una
                comp = conn.execute("SELECT immagine FROM componenti WHERE id=?", (cid,)).fetchone()
                if comp:
                    new_img = comp[0] or url
                    conn.execute("UPDATE componenti SET immagine=?, aggiornato_il=datetime('now') WHERE id=?",
                                 (new_img, cid))
                conn.commit()
            finally:
                conn.close()
        # Notifica il PC in tempo reale
        try:
            socketio.emit("inventario_foto", {
                "sessione_id": riga["sessione_id"],
                "riga_id": rid,
                "componente_id": cid,
                "url": url,
                "foto_url": new_foto_url
            }, room=f"inventario_{riga['sessione_id']}")
        except: pass
        return jsonify({"ok": True, "url": url})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/inventario/sessioni/<int:sid>/chiudi", methods=["POST"])
@require_login
def api_inv_chiudi(sid):
    try:
        with db._write_lock:
            conn = db.get_connection()
            try:
                # Sincronizza immagini inventario → componenti
                righe = conn.execute(
                    """SELECT componente_id, foto_url FROM inventario_righe
                       WHERE sessione_id=? AND foto_url IS NOT NULL AND foto_url != ''""",
                    (sid,)
                ).fetchall()
                for r in righe:
                    prima_foto = r["foto_url"].split("|")[0].strip()
                    if prima_foto:
                        conn.execute(
                            "UPDATE componenti SET immagine=?, aggiornato_il=datetime('now') WHERE id=?",
                            (prima_foto, r["componente_id"])
                        )
                conn.execute(
                    "UPDATE sessioni_inventario SET stato='chiusa', chiuso_il=datetime('now') WHERE id=?",
                    (sid,)
                )
                conn.commit()
            except Exception as e:
                conn.rollback(); raise
            finally:
                conn.close()
        _comp_cache["ts"] = 0
        socketio.emit("inventario_chiusa", {"sessione_id": sid})
        socketio.emit("magazzino_aggiornato", {"reload": True})
        return jsonify({"ok": True, "msg": "Sessione chiusa"})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

# SocketIO room per inventario

@socketio.on("join_inventario")
def on_join_inventario(data):
    sid = data.get("sessione_id")
    if sid:
        join_room(f"inventario_{sid}")
        log.info(f"Client joined inventario_{sid}")

@socketio.on("leave_inventario")
def on_leave_inventario(data):
    sid = data.get("sessione_id")
    if sid:
        leave_room(f"inventario_{sid}")

@socketio.on("sync_completata")
def on_sync_completata(data):
    """Telefono ha sincronizzato operazioni offline — notifica il PC."""
    sid = data.get("sessione_id")
    n   = data.get("n", 0)
    if sid:
        socketio.emit("inventario_sync_completata", {"sessione_id": sid, "n": n},
                      room=f"inventario_{sid}")

@app.route("/api/ping")
def api_ping():
    """Endpoint pubblico leggero per verificare connessione al server.
    Non espone dati sensibili — solo conferma che il server è raggiungibile.
    """
    return jsonify({"ok": True})

@app.route("/sw.js")
def service_worker():
    from flask import send_from_directory
    resp = send_from_directory("web/static", "sw.js")
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

# ══ IMPORT EXCEL BACKGROUND ══════════════════════════════════════════
_import_stato = {"running": False, "importati": 0, "aggiornati": 0,
                  "saltati": 0, "totale": 0, "processed": 0, "msg": "", "ok": None}

def _processa_import_bg(tmp_path, utente_id, utente_username):
    """Processa import Excel in background thread."""
    global _import_stato
    import openpyxl, os as _os

    _import_stato.update({"running": True, "importati": 0, "aggiornati": 0,
                           "saltati": 0, "totale": 0, "processed": 0,
                           "msg": "Avvio importazione...", "ok": None})
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(h or "").strip().lower() for h in header_row]
        col_map = {}
        for i, h in enumerate(header):
            field = DANEA_MAP.get(h, h.replace(" ","_").replace(".","").replace("'",""))
            col_map[field] = i

        log.info(f"Import BG: colonne trovate: {list(col_map.keys())}")

        def get(row, field, default=None):
            idx = col_map.get(field)
            if idx is None or idx >= len(row): return default
            v = row[idx]
            if v is None: return default
            import datetime
            if isinstance(v, (datetime.datetime, datetime.date)):
                return str(v.year)
            s = str(v).strip()
            return s if s not in ("None","nan","") else default

        def toint(row, field):
            try: v = get(row, field, "0"); return int(float(v)) if v else 0
            except: return 0

        importati = 0; aggiornati = 0; saltati = 0; row_n = 0
        BATCH = 500

        with db._write_lock:
            conn = db.get_connection()
            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_n += 1
                    codice = get(row, "codice") or get(row, "cmp")
                    nome   = get(row, "nome") or get(row, "articolo") or get(row, "descrizione")

                    if not codice and not nome:
                        saltati += 1; continue
                    if not codice: codice = "IMP-" + str(row_n).zfill(5)
                    if not nome:   nome   = codice

                    campi = {
                        _DB_COLS.get("articolo","nome"): nome,
                        "tipologia": get(row,"tipologia"),
                        "categoria": get(row,"categoria"),
                        "sottocategoria": get(row,"sottocategoria"),
                        "cod_udm": get(row,"cod_udm") or "PZ",
                        "cod_iva": get(row,"cod_iva"),
                        "listino1": get(row,"listino1"),
                        "listino2": get(row,"listino2"),
                        "listino3": get(row,"listino3"),
                        "note": get(row,"nota") or get(row,"note"),
                        "cod_barre": get(row,"cod_barre"),
                        "produttore": get(row,"produttore") or get(row,"marca"),
                        "modello": get(row,"modello"),
                        "cod_fornitore": get(row,"cod_fornitore"),
                        "fornitore": get(row,"fornitore"),
                        "ubicazione": get(row,"ubicazione"),
                        "colore": get(row,"colore"),
                        "anno": get(row,"anno") or get(row,"anno_da"),
                        "alimentazione": get(row,"alimentazione") or get(row,"carburante"),
                        "extra3": get(row,"extra3"),
                        "cod_prod_forn": get(row,"cod_prod_forn"),
                        "cilindrata": get(row,"cilindrata"),
                    }
                    valid_cols = set(_DB_COLS.get("all_cols", []))
                    campi = {k: v for k, v in campi.items()
                             if v is not None and (not valid_cols or k in valid_cols)}
                    esistenza = toint(row, "esistenza") or toint(row, "quantita") or toint(row, "giacenza")

                    existing = conn.execute(
                        f"SELECT id FROM componenti WHERE {_DB_COLS.get('cmp','codice')}=? AND {_DB_COLS.get('eliminato','eliminato')}=0", (codice,)
                    ).fetchone()

                    if existing:
                        comp_id = existing[0]
                        campi_update = {k:v for k,v in campi.items() if v is not None and v != ""}
                        if campi_update:
                            sets = ", ".join(f"{k}=?" for k in campi_update)
                            conn.execute(
                                f"UPDATE componenti SET {sets}, {_DB_COLS.get('aggiornato_il','aggiornato_il')}=datetime('now') WHERE id=?",
                                list(campi_update.values()) + [comp_id])
                        aggiornati += 1
                    else:
                        campi_ins = {k:v for k,v in campi.items() if v is not None and v != ""}
                        campi_ins[_DB_COLS.get("cmp","codice")]      = codice
                        campi_ins["eliminato"] = 0
                        cols_str = ", ".join(campi_ins.keys())
                        phs = ", ".join(["?"] * len(campi_ins))
                        cur = conn.execute(
                            f"INSERT INTO componenti ({cols_str}) VALUES ({phs})",
                            list(campi_ins.values()))
                        comp_id = cur.lastrowid
                        importati += 1

                    if esistenza > 0:
                        # Prova tabella movimenti con schema flessibile
                        try:
                            already = conn.execute(
                                f"SELECT id FROM {_DB_COLS.get('tabella_movimenti','movimenti')} WHERE componente_id=? AND riferimento='Import Excel' LIMIT 1",
                                (comp_id,)).fetchone()
                            if not already:
                                mov_cols = [r[1] for r in conn.execute("PRAGMA table_info(movimenti)").fetchall()]
                                if "username" in mov_cols:
                                    conn.execute(
                                        f"INSERT INTO {_DB_COLS.get('tabella_movimenti','movimenti')} (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento,username) VALUES(?,?,?,?,?,?,?)",
                                        (comp_id,"carico",esistenza,0,esistenza,"Import Excel",utente_username))
                                else:
                                    conn.execute(
                                        f"INSERT INTO {_DB_COLS.get('tabella_movimenti','movimenti')} (componente_id,tipo,quantita,quantita_prima,quantita_dopo,riferimento) VALUES(?,?,?,?,?,?)",
                                        (comp_id,"carico",esistenza,0,esistenza,"Import Excel"))
                        except Exception as _me:
                            log.warning(f"Mov insert skip: {_me}")

                    if row_n % BATCH == 0:
                        conn.commit()
                        _import_stato.update({
                            "importati": importati, "aggiornati": aggiornati,
                            "saltati": saltati, "processed": row_n,
                            "msg": f"⏳ {importati} nuovi, {aggiornati} aggiornati, riga {row_n}..."
                        })
                        socketio.emit("import_progress", {
                            "importati": importati, "aggiornati": aggiornati,
                            "processed": row_n, "saltati": saltati
                        })

                conn.commit()
            finally:
                conn.close()

        try: _os.unlink(tmp_path)
        except: pass

        msg = f"✅ Import completato: {importati} nuovi, {aggiornati} aggiornati, {saltati} saltati su {row_n} righe"
        log.info(msg)
        _import_stato.update({"running": False, "importati": importati, "aggiornati": aggiornati,
                               "saltati": saltati, "processed": row_n, "msg": msg, "ok": True})
        socketio.emit("import_done", {"ok": True, "msg": msg,
                                       "importati": importati, "aggiornati": aggiornati, "saltati": saltati})

    except Exception as e:
        log.exception(f"Import BG errore: {e}")
        msg = f"❌ Errore: {e}"
        _import_stato.update({"running": False, "msg": msg, "ok": False})
        socketio.emit("import_done", {"ok": False, "msg": msg})
        try: _os.unlink(tmp_path)
        except: pass


@app.route("/api/magazzino/import-excel-bg", methods=["POST"])
@require_login
def api_import_excel_bg():
    """Avvia import Excel in background — ritorna subito, progresso via SocketIO."""
    global _import_stato
    if _import_stato.get("running"):
        return jsonify({"ok": False, "msg": "Import già in corso"}), 400

    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "Nessun file"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx",".xls")):
        return jsonify({"ok": False, "msg": "Solo .xlsx o .xls"}), 400

    import os as _os
    u = cu()
    tmp_path = str(ROOT / "logs" / "_import_tmp.xlsx")
    f.save(tmp_path)
    log.info(f"Import BG: file salvato {_os.path.getsize(tmp_path)//1024}KB")

    t = threading.Thread(target=_processa_import_bg,
                         args=(tmp_path, u.get("id"), u.get("username")),
                         daemon=True)
    t.start()
    return jsonify({"ok": True, "msg": "Import avviato in background"})


@app.route("/api/magazzino/import-stato")
@require_login
def api_import_stato():
    return jsonify(_import_stato)


# ══ AGGIORNAMENTI ════════════════════════════════════════════════════
VERSIONE_APP   = "3.6.3"        # aggiornato da pubblica_aggiornamento.py
_PROD_REPO     = "00mine/perilcar-prod"
_PROD_BRANCH   = "main"

_update_cache      = None
_update_checked_at = 0.0


def _gh_token() -> str:
    """Legge il Personal Access Token GitHub (sola lettura) da config/settings.json."""
    return cfg.get("github_token", "")


def _controlla_aggiornamenti():
    """Controlla il repo di produzione per nuova versione. Silenzioso, thread-safe."""
    global _update_cache, _update_checked_at
    import urllib.request, json as _json
    try:
        token = _gh_token()
        api_url = (
            f"https://api.github.com/repos/{_PROD_REPO}/contents/version.json"
            f"?ref={_PROD_BRANCH}"
        )
        req = urllib.request.Request(api_url)
        req.add_header("User-Agent",  f"PerilCar-ERP/{VERSIONE_APP}")
        req.add_header("Accept",      "application/vnd.github.v3.raw")
        if token:
            req.add_header("Authorization", f"Bearer {token}")
        with urllib.request.urlopen(req, timeout=8) as r:
            data = _json.loads(r.read().decode())
        remota = data.get("versione", "0.0.0")
        rv = tuple(int(x) for x in remota.split("."))
        lv = tuple(int(x) for x in VERSIONE_APP.split("."))
        _update_cache = data if rv > lv else None
        _update_checked_at = time.time()
        if _update_cache:
            log.info(f"Aggiornamento disponibile: v{remota}")
    except Exception as e:
        log.debug(f"Check aggiornamenti: {e}")


def _avvia_check_aggiornamenti():
    """Avvia il check in background; si ripete ogni 6 ore."""
    def _loop():
        while True:
            _controlla_aggiornamenti()
            time.sleep(6 * 3600)
    threading.Thread(target=_loop, daemon=True, name="check-update").start()


_avvia_check_aggiornamenti()


@app.route("/api/check-update")
@require_login
def api_check_update():
    if _update_cache:
        return jsonify({
            "disponibile":  True,
            "versione":     _update_cache.get("versione"),
            "changelog":    _update_cache.get("changelog", []),
            "obbligatorio": _update_cache.get("obbligatorio", False),
            "data":         _update_cache.get("data", ""),
            "corrente":     VERSIONE_APP,
        })
    return jsonify({"disponibile": False, "corrente": VERSIONE_APP})


def _riavvia_server():
    """Riavvia il processo server dopo 3 sec tramite script batch esterno (funziona sia come .py che .exe)."""
    import tempfile, subprocess as _sp
    if getattr(sys, "frozen", False):
        cmd = f'"{sys.executable}"'
    else:
        cmd = f'"{sys.executable}" "{Path(__file__).absolute()}"'
    bat = Path(tempfile.gettempdir()) / "perilcar_restart.bat"
    bat.write_text(
        f"@echo off\ntimeout /t 3 /nobreak >nul\nstart \"PerilCar\" {cmd}\n",
        encoding="utf-8",
    )
    _sp.Popen(
        ["cmd", "/c", str(bat)],
        creationflags=_sp.DETACHED_PROCESS | _sp.CREATE_NEW_CONSOLE,
        close_fds=True,
    )
    time.sleep(0.5)
    os._exit(0)


@app.route("/api/installa-aggiornamento", methods=["POST"])
@require_login
def api_installa_aggiornamento():
    """Scarica il ZIP dal repo di produzione e aggiorna i file. Solo admin."""
    import urllib.request, zipfile, shutil

    u = session.get("user", {})
    if u.get("ruolo") != "admin":
        return jsonify({"ok": False, "msg": "Solo gli amministratori possono installare aggiornamenti."}), 403

    token = _gh_token()
    if not token:
        return jsonify({
            "ok":  False,
            "msg": "Token GitHub non configurato in config/settings.json. "
                   "Contattare l'amministratore di sistema.",
        }), 500

    tmp_zip = ROOT / "logs" / "_update.zip"
    try:
        zip_url = f"https://api.github.com/repos/{_PROD_REPO}/zipball/{_PROD_BRANCH}"
        ver_nuova = (_update_cache or {}).get("versione", "?")
        log.info(f"Download aggiornamento v{ver_nuova} da repo prod")

        req = urllib.request.Request(zip_url)
        req.add_header("User-Agent",    f"PerilCar-ERP/{VERSIONE_APP}")
        req.add_header("Accept",        "application/vnd.github+json")
        req.add_header("Authorization", f"Bearer {token}")
        with urllib.request.urlopen(req, timeout=120) as r:
            with open(tmp_zip, "wb") as f:
                shutil.copyfileobj(r, f)
        log.info(f"ZIP scaricato: {tmp_zip.stat().st_size:,} byte")

        # Cartelle e file da NON toccare durante un aggiornamento
        SKIP_DIRS  = {"db", "backup", "logs", "config", "web/static/uploads"}
        SKIP_FILES = {"config/settings.json", ".gitignore"}
        SKIP_EXT   = {".db", ".db-wal", ".db-shm", ".log", ".bak"}

        with zipfile.ZipFile(tmp_zip) as zf:
            names  = zf.namelist()
            prefix = names[0].split("/")[0] + "/"
            aggiornati = 0
            for member in names:
                if member == prefix:
                    continue
                rel = member[len(prefix):]
                if not rel:
                    continue
                if any(rel == d or rel.startswith(d + "/") for d in SKIP_DIRS):
                    continue
                if rel in SKIP_FILES:
                    continue
                if Path(rel).suffix.lower() in SKIP_EXT:
                    continue
                dest = ROOT / rel
                if member.endswith("/"):
                    dest.mkdir(parents=True, exist_ok=True)
                else:
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    with zf.open(member) as src_f, open(dest, "wb") as dst_f:
                        dst_f.write(src_f.read())
                    aggiornati += 1

        tmp_zip.unlink(missing_ok=True)
        log.info(f"Aggiornamento installato: {aggiornati} file aggiornati")
        db.log(u.get("id"), u.get("username"), "sistema", "aggiornamento_installato",
               dati_nuovi=str({"versione": ver_nuova}))

        threading.Thread(target=_riavvia_server, daemon=True).start()
        return jsonify({"ok": True, "msg": "Aggiornamento installato. Riavvio in corso..."})

    except Exception as e:
        log.exception(f"Errore installazione aggiornamento: {e}")
        tmp_zip.unlink(missing_ok=True)
        return jsonify({"ok": False, "msg": f"Errore: {str(e)}"}), 500

# ══ STATO SISTEMA ════════════════════════════════════════════════════
VERSIONE_APP = "3.6.3"

# Tracking utenti connessi (sid -> username)
_utenti_online_lock = threading.Lock()
_utenti_online = {}

@socketio.on("connect")
def _on_connect():
    try:
        from flask import request as _req
        u = session.get("user") or {}
        if u.get("username"):
            with _utenti_online_lock:
                _utenti_online[_req.sid] = {
                    "username": u.get("username"),
                    "nome":     u.get("nome", ""),
                    "from":     time.time(),
                }
    except Exception:
        pass

@socketio.on("disconnect")
def _on_disconnect():
    try:
        from flask import request as _req
        with _utenti_online_lock:
            _utenti_online.pop(_req.sid, None)
    except Exception:
        pass

def _get_server_ip():
    """Ritorna l'IP LAN del server."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def _get_ultimo_backup():
    """Cerca l'ultimo file di backup nella cartella."""
    bdir = ROOT / "backup"
    if not bdir.exists():
        return None
    files = sorted(bdir.glob("*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return None
    from datetime import datetime
    ts = datetime.fromtimestamp(files[0].stat().st_mtime)
    return ts.strftime("%d/%m/%Y %H:%M")

@app.route("/api/system-status")
@require_login
def api_system_status():
    import socket as _sock
    with _utenti_online_lock:
        utenti = [{"username": v["username"], "nome": v.get("nome", "")}
                  for v in _utenti_online.values()]
    # Deduplica per username (uno stesso utente può avere più tab)
    seen = set()
    utenti_unique = []
    for u in utenti:
        if u["username"] not in seen:
            seen.add(u["username"])
            utenti_unique.append(u)
    return jsonify({
        "online":         True,
        "ip":             _get_server_ip(),
        "porta":          request.host.split(":")[-1] if ":" in request.host else "80",
        "versione":       VERSIONE_APP,
        "utenti_online":  utenti_unique,
        "ultimo_backup":  _get_ultimo_backup(),
    })

# ══ VOCI TENDINE — voci personalizzabili nelle dropdown ═════════════════════
@app.route("/api/voci-tendine/<categoria>")
@require_login
def api_voci_get(categoria):
    rows = dem.all("SELECT id,valore FROM voci_tendine WHERE categoria=? AND valore!='' ORDER BY ordine,valore", (categoria,))
    return jsonify(rows)

@app.route("/api/voci-tendine/<categoria>", methods=["POST"])
@require_login
def api_voci_add(categoria):
    d = request.json or {}
    valore = (d.get("valore") or "").strip()
    if not valore:
        return jsonify({"ok": False, "msg": "Valore obbligatorio"}), 400
    try:
        dem.run("INSERT OR IGNORE INTO voci_tendine(categoria,valore) VALUES(?,?)", (categoria, valore))
        row = dem.one("SELECT id FROM voci_tendine WHERE categoria=? AND valore=?", (categoria, valore))
        return jsonify({"ok": True, "id": row["id"] if row else None, "valore": valore})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500

@app.route("/api/voci-tendine/<int:vid>", methods=["DELETE"])
@require_login
def api_voci_del(vid):
    dem.run("DELETE FROM voci_tendine WHERE id=?", (vid,))
    return jsonify({"ok": True})

@app.route("/api/magazzino/fix-view")
@require_login  
def api_fix_view():
    """Ricrea v_giacenza con le colonne reali del DB."""
    try:
        with db._write_lock:
            conn = db.get_connection()
            # Leggi colonne reali della tabella componenti
            cols = [r[1] for r in conn.execute("PRAGMA table_info(componenti)").fetchall()]
            log.info(f"Colonne componenti reali: {cols}")
            
            # Ricrea view con schema Danea (18 col)
            conn.execute("DROP VIEW IF EXISTS v_giacenza")
            conn.execute("""
                CREATE VIEW v_giacenza AS
                SELECT
                    c.id              AS componente_id,
                    c.codice          AS cmp,
                    c.descrizione     AS articolo,
                    c.produttore      AS marca,
                    c.modello,
                    c.udm             AS cod_udm,
                    c.anno            AS anno_da,
                    c.cod_prod_forn,
                    c.alimentazione   AS carburante,
                    c.colore,
                    c.note            AS nota,
                    c.cilindrata,
                    c.ubicazione,
                    c.cod_barre,
                    c.extra3,
                    c.cod_fornitore,
                    c.fornitore,
                    c.immagine        AS immagine_path,
                    c.eliminato,
                    c.aggiornato_il,
                    COALESCE(SUM(CASE
                        WHEN m.tipo IN ('carico','inventario') THEN  m.quantita
                        WHEN m.tipo = 'scarico'               THEN -m.quantita
                        WHEN m.tipo = 'rettifica'             THEN  m.quantita
                        ELSE 0
                    END), 0) AS esistenza
                FROM componenti c
                LEFT JOIN movimenti_magazzino m ON m.componente_id = c.id
                WHERE c.eliminato = 0
                GROUP BY c.id
            """)
            conn.commit()
            conn.close()
            log.info("v_giacenza ricreata con colonne corrette")
            return jsonify({"ok": True, "msg": f"View ricreata. Colonne: cmp={col_cmp}, articolo={col_art}"})
    except Exception as e:
        log.exception(f"Fix view error: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 500

# HOT RELOAD
# ══════════════════════════════════════════════════════════════════════════════

def start_file_watcher():
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
        class H(FileSystemEventHandler):
            def __init__(self): self._last = 0
            def on_modified(self, ev):
                if ev.is_directory: return
                if not any(ev.src_path.endswith(e) for e in (".py",".html",".css",".js")): return
                now = time.time()
                if now - self._last < 0.8: return
                self._last = now
                socketio.emit("reload", {"file": os.path.relpath(ev.src_path, ROOT)})
        obs = Observer(); obs.schedule(H(), str(ROOT), recursive=True); obs.start()
        log.info("👀 Watchdog attivo")
    except Exception as e:
        log.warning(f"Watchdog non disponibile: {e}")

if __name__ == "__main__":
    import webbrowser
    port = int(os.environ.get("PORT", 5000))
    threading.Thread(target=lambda: (time.sleep(1.5), webbrowser.open(f"http://localhost:{port}")),
                     daemon=True).start()
    start_file_watcher()
    log.info(f"🚀 PerilCar Dev Server — http://localhost:{port}")
    socketio.run(app, host="0.0.0.0", port=port,
                 debug=False, use_reloader=False, allow_unsafe_werkzeug=True)
